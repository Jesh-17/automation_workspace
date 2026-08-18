
import os
import re
import pandas as pd
from openpyxl import load_workbook  # ensures openpyxl engine availability


# =========================================================
# Config: Matching strictness (matching only; output values unchanged)
# =========================================================
# Options:
#   'exact'      -> no changes
#   'trim'       -> strip leading/trailing whitespace (DEFAULT)
#   'icase_trim' -> case-insensitive + strip whitespace
MATCH_MODE = 'trim'


# =========================================================
# Helpers: IO, Validation, and Column Standardization
# =========================================================

def ensure_dir(path: str):
    if path:
        os.makedirs(path, exist_ok=True)

def safe_read_excel(path: str, ctx: str = '') -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"[{ctx}] File not found: {path}")
    # Keep natural types; do not force dtype=str to respect raw values.
    return pd.read_excel(path, engine='openpyxl')

def safe_to_excel(df: pd.DataFrame, path: str):
    ensure_dir(os.path.dirname(path))
    df.to_excel(path, index=False, engine='openpyxl')

def require_columns(df: pd.DataFrame, required_cols: list, context: str = ""):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns {missing} in {context or 'DataFrame'}.")

# Standardize column names ONLY (values remain untouched)
COL_ALIASES = {
    'alarm description': 'AlarmDescription',
    'alarmdescription': 'AlarmDescription',
    'alarm_desc': 'AlarmDescription',
    'alarm desc': 'AlarmDescription',

    'priority': 'Priority',

    'ticketnumber': 'Ticketnumber',
    'ticket number': 'Ticketnumber',

    'date': 'Date',
    'time': 'Time',
    's.no': 'S.No',
    's no': 'S.No'
}

def _normalize_colname(name: str) -> str:
    key = re.sub(r'[^a-z0-9]', ' ', str(name).lower()).strip()
    key = re.sub(r'\s+', ' ', key)
    return COL_ALIASES.get(key, name)

def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {c: _normalize_colname(c) for c in df.columns}
    return df.rename(columns=mapping)


# =========================================================
# Part-1: Separation with/without tickets & Date/Time formatting
# =========================================================

def format_date_time_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if 'Date' in df.columns:
        def format_date(x):
            if pd.isna(x) or (isinstance(x, str) and x.strip() == ''):
                return ''
            dt = pd.to_datetime(x, errors='coerce', dayfirst=False)  # set dayfirst=True if your source is D/M/Y
            return dt.strftime('%m/%d/%Y') if pd.notna(dt) else str(x)
        df['Date'] = df['Date'].apply(format_date)

    if 'Time' in df.columns:
        def format_time(x):
            if pd.isna(x) or (isinstance(x, str) and x.strip() == ''):
                return ''
            dt = pd.to_datetime(x, errors='coerce')
            return dt.strftime('%I:%M:%S %p') if pd.notna(dt) else str(x)
        df['Time'] = df['Time'].apply(format_time)

    return df

_EMPTY_TOKENS = {'', 'nan', 'none', 'null'}

def _is_blank_ticket(val) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    s = str(val).strip()
    return s.lower() in _EMPTY_TOKENS

def split_by_ticket_presence(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    require_columns(df, ['Ticketnumber'], context='split_by_ticket_presence')
    mask_blank = df['Ticketnumber'].apply(_is_blank_ticket)
    without_tickets = df[mask_blank].copy()
    with_tickets = df[~mask_blank].copy()
    return with_tickets, without_tickets

def reset_serial_numbers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df['S.No'] = range(1, len(df) + 1)
    return df

def save_to_excel(df: pd.DataFrame, output_path: str):
    ensure_dir(os.path.dirname(output_path))
    if os.path.exists(output_path):
        os.remove(output_path)
    df.to_excel(output_path, index=False, engine='openpyxl')

def alarms_separation():
    input_path = './wsr/output/parent_filling_using_ticket_tracker/parent_filling_using_ticket_tracker_in_this_day_wise_tiggered_alarms_list.xlsx'
    output_file_with_tickets = './tickets_separation/output/alarms_having_tickets.xlsx'
    output_file_without_tickets = './tickets_separation/output/alarms_not_having_tickets.xlsx'

    try:
        df = safe_read_excel(input_path, ctx='alarms_separation')
    except FileNotFoundError:
        print(f"File not found: {input_path}")
        return

    df = standardize_columns(df)
    df = format_date_time_columns(df)
    require_columns(df, ['Ticketnumber'], context='alarms_separation')

    with_tickets, without_tickets = split_by_ticket_presence(df)
    with_tickets = reset_serial_numbers(with_tickets)
    without_tickets = reset_serial_numbers(without_tickets)

    save_to_excel(with_tickets, output_file_with_tickets)
    save_to_excel(without_tickets, output_file_without_tickets)

    print("Files created successfully:")
    print(f"- {output_file_with_tickets}")
    print(f"- {output_file_without_tickets}")


# =========================================================
# Part-2: Custom sheet creation & Aggregations with ROW-LEVEL matching
# =========================================================

def create_custom_columns_excel():
    input_path = './tickets_separation/output/alarms_not_having_tickets.xlsx'
    output_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'

    df = safe_read_excel(input_path, ctx='create_custom_columns_excel')
    df = standardize_columns(df)
    require_columns(df, ['Date', 'AlarmDescription', 'Priority'], context='create_custom_columns_excel')

    custom_df = pd.DataFrame({
        'Date': df['Date'],
        'No Of Days Trigerred': [''] * len(df),  # keep spelling as you used
        'AlarmDescription': df['AlarmDescription'],
        'Priority': df['Priority'],
        'Count of AlarmDescription': [''] * len(df)
    })
    safe_to_excel(custom_df, output_file)
    print(f"Custom Excel file created successfully at: {output_file}")


def _format_dates_smart(dates_like: list) -> str:
    """
    Aggregate dates into:
      - One date:           'Jan-12'
      - Same month:         'Jan-12, 13, 14, 15'
      - Different months:   'Jan-12, Jan-16'
    """
    if not dates_like:
        return ''
    ser = pd.to_datetime(pd.Series(dates_like), errors='coerce')
    ser = ser.dropna().sort_values()
    if ser.empty:
        return ''

    first = ser.iloc[0]
    same_month = (ser.dt.month == first.month) & (ser.dt.year == first.year)

    if same_month.all():
        parts = ser.dt.strftime('%b-%d').tolist()
        if len(parts) == 1:
            return parts[0]
        first_part = parts[0]                     # e.g., 'Jan-12'
        day_parts = [p.split('-')[1] for p in parts[1:]]  # ['13','14','15']
        return f"{first_part}, {', '.join(day_parts)}"
    else:
        # Different month/year → list all fully
        return ', '.join(ser.dt.strftime('%b-%d').tolist())


# ---------- Matching (row-level) ----------

def _norm(v):
    """Normalize keys for matching only, depending on MATCH_MODE."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        s = ''
    else:
        s = str(v)
    if MATCH_MODE == 'exact':
        return s
    s = s.strip()  # trim mode and icase_trim
    if MATCH_MODE == 'icase_trim':
        return s.casefold()
    return s  # 'trim'

def _build_ref_keysets(ref_df: pd.DataFrame):
    """
    Build two sets from the reference:
      - desc_only: AlarmDescription where Priority is blank/missing
      - desc_with_pri: tuples (AlarmDescription, Priority) where Priority is present
    Row-level logic avoids over-constraining matches when some ref rows have blank Priority.
    """
    require_columns(ref_df, ['AlarmDescription'], context='reference')
    desc_col = ref_df['AlarmDescription']
    pri_col = ref_df['Priority'] if 'Priority' in ref_df.columns else pd.Series([None]*len(ref_df))

    desc_only = set()
    desc_with_pri = set()

    for d, p in zip(desc_col, pri_col):
        nd, np = _norm(d), _norm(p)
        if np == '':
            desc_only.add(nd)
        else:
            desc_with_pri.add((nd, np))
    return desc_only, desc_with_pri

def _build_custom_keys(df: pd.DataFrame):
    require_columns(df, ['AlarmDescription', 'Priority'], context='custom_df')
    k_desc = df['AlarmDescription'].map(_norm)
    k_pri = df['Priority'].map(_norm)
    return k_desc, k_pri

def _mask_semijoin_rowlevel(custom_df: pd.DataFrame, ref_df: pd.DataFrame) -> pd.Series:
    """
    For each custom row, match if:
      - custom.desc in ref.desc_only  OR
      - (custom.desc, custom.priority) in ref.desc_with_pri
    """
    desc_only, desc_with_pri = _build_ref_keysets(ref_df)
    k_desc, k_pri = _build_custom_keys(custom_df)
    return k_desc.isin(desc_only) | pd.Series(list(zip(k_desc, k_pri))).isin(desc_with_pri)

def _mask_antijoin_rowlevel(custom_df: pd.DataFrame, ref_df: pd.DataFrame) -> pd.Series:
    return ~_mask_semijoin_rowlevel(custom_df, ref_df)


# ---------- Aggregation writer (group by trimmed keys to avoid duplicates) ----------

def _aggregate_sheet_from_df(filtered_df: pd.DataFrame,
                             sheet_name: str,
                             custom_file: str,
                             priority_order: list[str] | None = None):
    """
    Aggregate by AlarmDescription + Priority using TRIMMED keys to avoid duplicates
    caused by trailing spaces. Display a clean (trimmed) representative label.
    """
    if filtered_df.empty:
        out_df = pd.DataFrame(columns=['Date', 'No Of Days Trigerred', 'AlarmDescription', 'Priority', 'Count of AlarmDescription'])
    else:
        # Build grouping keys (trim only — we are NOT changing reference matching)
        df = filtered_df.copy()
        df['__desc_key'] = df['AlarmDescription'].astype(str).str.strip()
        df['__pri_key']  = df['Priority'].astype(str).str.strip()

        # Representative label (trimmed)
        def _repr_label(series: pd.Series) -> str:
            trimmed = series.astype(str).str.strip()
            if trimmed.empty:
                return ''
            modes = trimmed.mode(dropna=False)
            return modes.iloc[0] if not modes.empty else trimmed.iloc[0]

        rows = []
        for (dkey, pkey), g in df.groupby(['__desc_key', '__pri_key'], dropna=False):
            # Representative labels (display)
            rep_desc = _repr_label(g['AlarmDescription'])
            rep_prio = _repr_label(g['Priority'])

            # Unique dates, then pretty-format
            dates = [d for d in g['Date'].tolist() if isinstance(d, str) and d.strip() != '']
            unique_dates = sorted(set(dates))
            date_str = _format_dates_smart(unique_dates)

            rows.append({
                'Date': date_str,
                'No Of Days Trigerred': len(unique_dates),
                'AlarmDescription': rep_desc,
                'Priority': rep_prio,
                'Count of AlarmDescription': len(g)
            })

        out_df = pd.DataFrame(rows)

    # Sort as before
    if not out_df.empty:
        if priority_order:
            out_df['Priority'] = pd.Categorical(out_df['Priority'], categories=priority_order, ordered=True)
            out_df = out_df.sort_values(by=['Priority', 'No Of Days Trigerred', 'Count of AlarmDescription'],
                                        ascending=[True, False, False])
        else:
            out_df = out_df.sort_values(by=['Priority', 'No Of Days Trigerred', 'Count of AlarmDescription'],
                                        ascending=[True, False, False])

    # Write/replace
    ensure_dir(os.path.dirname(custom_file))
    with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        out_df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Sheet '{sheet_name}' added successfully. Rows: {len(out_df)}")


# ---------- Sheet builders using row-level matching ----------

def add_duration_alarms_sheet(priority_order: list[str] | None = None):
    custom_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'
    duration_file = './required_files/duration_alarms_list.xlsx'

    custom_df = safe_read_excel(custom_file, ctx='duration:custom')
    custom_df = standardize_columns(custom_df)

    duration_df = safe_read_excel(duration_file, ctx='duration:reference')
    duration_df = standardize_columns(duration_df)

    require_columns(custom_df, ['AlarmDescription', 'Priority', 'Date'], context='duration:custom')
    require_columns(duration_df, ['AlarmDescription'], context='duration:reference')

    mask = _mask_semijoin_rowlevel(custom_df, duration_df)
    filtered = custom_df[mask].copy()
    print(f"[duration] matched: {mask.sum()} / {len(custom_df)}")

    _aggregate_sheet_from_df(filtered, sheet_name='duration_alarms', custom_file=custom_file, priority_order=priority_order)


def add_re_fine_tuning_alarms_sheet(priority_order: list[str] | None = None):
    custom_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'
    re_fine_file = './required_files/re_fine_tuning_alarms_list.xlsx'

    custom_df = safe_read_excel(custom_file, ctx='re_fine:custom')
    custom_df = standardize_columns(custom_df)

    re_fine_df = safe_read_excel(re_fine_file, ctx='re_fine:reference')
    re_fine_df = standardize_columns(re_fine_df)

    require_columns(custom_df, ['AlarmDescription', 'Priority', 'Date'], context='re_fine:custom')
    require_columns(re_fine_df, ['AlarmDescription'], context='re_fine:reference')

    mask = _mask_semijoin_rowlevel(custom_df, re_fine_df)
    filtered = custom_df[mask].copy()
    print(f"[re_fine] matched: {mask.sum()} / {len(custom_df)}")

    _aggregate_sheet_from_df(filtered, sheet_name='re_fine_tuning_alarms', custom_file=custom_file, priority_order=priority_order)


def add_alarms_required_tickets_sheet(priority_order: list[str] | None = None):
    """
    alarms_required_tickets = custom_df - duration_matches - re_fine_matches
    Tie-breaker: duration precedence (excluded first), then re_fine.
    """
    custom_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'
    duration_file = './required_files/duration_alarms_list.xlsx'
    re_fine_file = './required_files/re_fine_tuning_alarms_list.xlsx'

    custom_df = safe_read_excel(custom_file, ctx='required:custom')
    custom_df = standardize_columns(custom_df)

    duration_df = safe_read_excel(duration_file, ctx='required:duration_ref')
    duration_df = standardize_columns(duration_df)

    re_fine_df = safe_read_excel(re_fine_file, ctx='required:re_fine_ref')
    re_fine_df = standardize_columns(re_fine_df)

    require_columns(custom_df, ['AlarmDescription', 'Priority', 'Date'], context='required:custom')
    require_columns(duration_df, ['AlarmDescription'], context='required:duration_ref')
    require_columns(re_fine_df, ['AlarmDescription'], context='required:re_fine_ref')

    # Exclude duration first (duration precedence)
    mask_dur = _mask_semijoin_rowlevel(custom_df, duration_df)
    after_duration = custom_df[~mask_dur].copy()
    print(f"[required] excluded (duration): {mask_dur.sum()}")

    # Exclude re_fine from the remainder
    mask_refine = _mask_semijoin_rowlevel(after_duration, re_fine_df)
    remaining = after_duration[~mask_refine].copy()
    print(f"[required] excluded (re_fine): {mask_refine.sum()}")

    _aggregate_sheet_from_df(remaining, sheet_name='alarms_required_tickets', custom_file=custom_file, priority_order=priority_order)


# =========================================================
# Orchestration
# =========================================================

def run_all(priority_order: list[str] | None = None):
    """
    Pipeline:
      1) Separate alarms with/without tickets (robust blank Ticketnumber).
      2) Create custom columns file.
      3) Add sheets via row-level matching:
         - 'duration_alarms'         (ref rows with blank Priority match on description only)
         - 're_fine_tuning_alarms'   (same row-level logic)
         - 'alarms_required_tickets' (remainder; duration precedence)
    Matching uses 'MATCH_MODE' for keys, but output values are unchanged.
    """
    alarms_separation()
    create_custom_columns_excel()
    add_duration_alarms_sheet(priority_order=priority_order)
    add_re_fine_tuning_alarms_sheet(priority_order=priority_order)
    add_alarms_required_tickets_sheet(priority_order=priority_order)
    print("Pipeline completed.")


# # Optional: run when executing this file directly
# if __name__ == '__main__':
#     # Example: enforce semantic priority order if you have one
#     # run_all(priority_order=['Critical', 'High', 'Medium', 'Low'])
#     run_all()






















# import os
# from openpyxl import load_workbook
# import pandas as pd
# import html




# def load_excel_file(file_path):
#     try:
#         df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
#         return df
#     except FileNotFoundError:
#         print(f"File not found: {file_path}")
#         return None

# def format_date_time_columns(df):
#     df = df.copy()
#     # ✅ Dynamic Date formatting
#     if 'Date' in df.columns:
#         def format_date(x):
#             if pd.isna(x) or str(x).strip() == '':
#                 return None
#             try:
#                 dt = pd.to_datetime(x, errors='coerce')
#                 return dt.strftime('%m/%d/%Y') if pd.notna(dt) else str(x)
#             except Exception:
#                 return str(x)
#         df['Date'] = df['Date'].apply(format_date)

#     # ✅ Dynamic Time formatting
#     if 'Time' in df.columns:
#         def format_time(x):
#             if pd.isna(x) or str(x).strip() == '':
#                 return None
#             try:
#                 dt = pd.to_datetime(x, errors='coerce')
#                 return dt.strftime('%I:%M:%S %p') if pd.notna(dt) else str(x)
#             except Exception:
#                 return str(x)
#         df['Time'] = df['Time'].apply(format_time)

#     return df

# def split_by_ticket_presence(df):
#     with_tickets = df[df['Ticketnumber'].notna() & df['Ticketnumber'].str.strip().ne('')].copy()
#     without_tickets = df[df['Ticketnumber'].isna() | df['Ticketnumber'].str.strip().eq('')].copy()
#     return with_tickets, without_tickets

# def reset_serial_numbers(df):
#     df = df.copy()
#     df['S.No'] = range(1, len(df) + 1)
#     return df

# def delete_existing_file(file_path):
#     if os.path.exists(file_path):
#         os.remove(file_path)

# def save_to_excel(df, output_path):
#     delete_existing_file(output_path)
#     df.to_excel(output_path, index=False, engine='openpyxl')

# def alarms_separation():
#     # input_path = './wsr/output/child_or_parent_corresponding_date_filled_in_this_day_wise_tiggered_alarms_list.xlsx'
#     input_path = './wsr/output/parent_filling_using_ticket_tracker/parent_filling_using_ticket_tracker_in_this_day_wise_tiggered_alarms_list.xlsx'
#     output_file_with_tickets = './tickets_separation/output/alarms_having_tickets.xlsx'
#     output_file_without_tickets = './tickets_separation/output/alarms_not_having_tickets.xlsx'

#     # base_dir = os.path.dirname(__file__)  # gets the current alarms_processor.py path
#     # input_path = os.path.join(base_dir, 'required_files_for_tickets_separation', 'day_wise_tiggered_alarms_list.xlsx')
#     # output_file_with_tickets = os.path.join(base_dir, 'alarms_having_tickets.xlsx')
#     # output_file_without_tickets = os.path.join(base_dir, 'alarms_not_having_tickets.xlsx')

#     df = load_excel_file(input_path)

#     if df is not None:
#         df = format_date_time_columns(df)
#         with_tickets, without_tickets = split_by_ticket_presence(df)

#         with_tickets = reset_serial_numbers(with_tickets)
#         without_tickets = reset_serial_numbers(without_tickets)

#         save_to_excel(with_tickets, output_file_with_tickets)
#         save_to_excel(without_tickets, output_file_without_tickets)

#         print("Files created successfully:")
#         print(f"- {output_file_with_tickets}")
#         print(f"- {output_file_without_tickets}")



# def format_dates(dates):
#     """Format dates as 'Mon-DD,DD & DD'."""
#     formatted_dates = [pd.to_datetime(d).strftime('%b-%d') for d in dates]
#     if len(formatted_dates) > 1:
#         first_date = formatted_dates[0]
#         day_parts = [d.split('-')[1] for d in formatted_dates[1:]]
#         if len(day_parts) > 1:
#             return f"{first_date},{','.join(day_parts[:-1])}, {day_parts[-1]}"
#         else:
#             return f"{first_date}, {day_parts[0]}"
#     else:
#         return formatted_dates[0]


    
# def create_custom_columns_excel():
#     input_path = './tickets_separation/output/alarms_not_having_tickets.xlsx'
#     output_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'

#     df = pd.read_excel(input_path, engine='openpyxl')
#     if df is not None:
#         custom_df = pd.DataFrame({
#             'Date': df['Date'],
#             'No Of Days Trigerred': [''] * len(df),
#             'AlarmDescription': df['AlarmDescription'],
#             'Priority': df['Priority'],
#             'Count of AlarmDescription': [''] * len(df)
#         })
#         custom_df.to_excel(output_file, index=False)
#         print(f"Custom Excel file created successfully at: {output_file}")


# def add_duration_alarms_sheet():
#     custom_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'
#     duration_file = './required_files/duration_alarms_list.xlsx'

#     custom_df = pd.read_excel(custom_file, engine='openpyxl')
#     duration_df = pd.read_excel(duration_file, engine='openpyxl')

#     duration_alarms_set = set(duration_df['AlarmDescription'].dropna().unique())
#     filtered_df = custom_df[custom_df['AlarmDescription'].isin(duration_alarms_set)].copy()

#     grouped = filtered_df.groupby(['AlarmDescription', 'Priority'])
#     duration_data = []
#     for (alarm_desc, priority), group in grouped:
#         unique_dates = sorted(group['Date'].dropna().unique())
#         date_str = format_dates(unique_dates)
#         duration_data.append({
#             'Date': date_str,
#             'No Of Days Trigerred': len(unique_dates),
#             'AlarmDescription': alarm_desc,
#             'Priority': priority,
#             'Count of AlarmDescription': len(group)
#         })

#     duration_df_final = pd.DataFrame(duration_data)
#     duration_df_final.sort_values(by=['Priority', 'No Of Days Trigerred', 'Count of AlarmDescription'],
#                                   ascending=[True, False, False], inplace=True)

#     with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         duration_df_final.to_excel(writer, sheet_name='duration_alarms', index=False)

#     print("Sheet 'duration_alarms' added successfully.")


# def add_re_fine_tuning_alarms_sheet():
#     custom_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'
#     re_fine_tuning_file = './required_files/re_fine_tuning_alarms_list.xlsx'

#     custom_df = pd.read_excel(custom_file, engine='openpyxl')
#     re_fine_tuning_df = pd.read_excel(re_fine_tuning_file, engine='openpyxl')

#     re_fine_tuning_alarms_set = set(re_fine_tuning_df['AlarmDescription'].dropna().unique())
#     filtered_df = custom_df[custom_df['AlarmDescription'].isin(re_fine_tuning_alarms_set)].copy()

#     grouped = filtered_df.groupby(['AlarmDescription', 'Priority'])
#     re_fine_tuning_data = []
#     for (alarm_desc, priority), group in grouped:
#         unique_dates = sorted(group['Date'].dropna().unique())
#         date_str = format_dates(unique_dates)
#         re_fine_tuning_data.append({
#             'Date': date_str,
#             'No Of Days Trigerred': len(unique_dates),
#             'AlarmDescription': alarm_desc,
#             'Priority': priority,
#             'Count of AlarmDescription': len(group)
#         })

#     re_fine_tuning_df_final = pd.DataFrame(re_fine_tuning_data)
#     re_fine_tuning_df_final.sort_values(by=['Priority', 'No Of Days Trigerred', 'Count of AlarmDescription'],
#                                         ascending=[True, False, False], inplace=True)

#     with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         re_fine_tuning_df_final.to_excel(writer, sheet_name='re_fine_tuning_alarms', index=False)

#     print("Sheet 're_fine_tuning_alarms' added successfully.")


# def add_alarms_required_tickets_sheet():
#     custom_file = './tickets_separation/output/alarms_not_having_tickets_custom_columns.xlsx'
#     duration_file = './required_files/duration_alarms_list.xlsx'
#     re_fine_tuning_file = './required_files/re_fine_tuning_alarms_list.xlsx'

#     custom_df = pd.read_excel(custom_file, engine='openpyxl')
#     duration_df = pd.read_excel(duration_file, engine='openpyxl')
#     re_fine_tuning_df = pd.read_excel(re_fine_tuning_file, engine='openpyxl')

#     duration_alarms_set = set(duration_df['AlarmDescription'].dropna().unique())
#     re_fine_tuning_alarms_set = set(re_fine_tuning_df['AlarmDescription'].dropna().unique())

#     excluded_alarms = duration_alarms_set.union(re_fine_tuning_alarms_set)
#     remaining_df = custom_df[~custom_df['AlarmDescription'].isin(excluded_alarms)].copy()

#     grouped = remaining_df.groupby(['AlarmDescription', 'Priority'])
#     required_tickets_data = []
#     for (alarm_desc, priority), group in grouped:
#         unique_dates = sorted(group['Date'].dropna().unique())
#         date_str = format_dates(unique_dates)
#         required_tickets_data.append({
#             'Date': date_str,
#             'No Of Days Trigerred': len(unique_dates),
#             'AlarmDescription': alarm_desc,
#             'Priority': priority,
#             'Count of AlarmDescription': len(group)
#         })

#     required_tickets_df = pd.DataFrame(required_tickets_data)
#     required_tickets_df.sort_values(by=['Priority', 'No Of Days Trigerred', 'Count of AlarmDescription'],
#                                     ascending=[True, False, False], inplace=True)

#     with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         required_tickets_df.to_excel(writer, sheet_name='alarms_required_tickets', index=False)

#     print("Sheet 'alarms_required_tickets' added successfully.")









































































































































































































































































































































# import os
# from openpyxl import load_workbook
# import pandas as pd

# def load_excel_file(file_path):
#     try:
#         df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
#         return df
#     except FileNotFoundError:
#         print(f"File not found: {file_path}")
#         return None

# def format_date_time_columns(df):
#     df = df.copy()
#     # ✅ Dynamic Date formatting
#     if 'Date' in df.columns:
#         def format_date(x):
#             if pd.isna(x) or str(x).strip() == '':
#                 return None
#             try:
#                 dt = pd.to_datetime(x, errors='coerce')
#                 return dt.strftime('%m/%d/%Y') if pd.notna(dt) else str(x)
#             except Exception:
#                 return str(x)
#         df['Date'] = df['Date'].apply(format_date)

#     # ✅ Dynamic Time formatting
#     if 'Time' in df.columns:
#         def format_time(x):
#             if pd.isna(x) or str(x).strip() == '':
#                 return None
#             try:
#                 dt = pd.to_datetime(x, errors='coerce')
#                 return dt.strftime('%I:%M:%S %p') if pd.notna(dt) else str(x)
#             except Exception:
#                 return str(x)
#         df['Time'] = df['Time'].apply(format_time)

#     return df

# def split_by_ticket_presence(df):
#     with_tickets = df[df['Ticketnumber'].notna() & df['Ticketnumber'].str.strip().ne('')].copy()
#     without_tickets = df[df['Ticketnumber'].isna() | df['Ticketnumber'].str.strip().eq('')].copy()
#     return with_tickets, without_tickets

# def reset_serial_numbers(df):
#     df = df.copy()
#     df['S.No'] = range(1, len(df) + 1)
#     return df

# def delete_existing_file(file_path):
#     if os.path.exists(file_path):
#         os.remove(file_path)

# def save_to_excel(df, output_path):
#     delete_existing_file(output_path)
#     df.to_excel(output_path, index=False, engine='openpyxl')

# def alarms_separation():
#     input_path = './wsr/output/child_or_parent_corresponding_date_filled_in_this_day_wise_tiggered_alarms_list.xlsx'
#     output_file_with_tickets = './tickets_separation/alarms_having_tickets.xlsx'
#     output_file_without_tickets = './tickets_separation/alarms_not_having_tickets.xlsx'

#     # base_dir = os.path.dirname(__file__)  # gets the current alarms_processor.py path
#     # input_path = os.path.join(base_dir, 'required_files_for_tickets_separation', 'day_wise_tiggered_alarms_list.xlsx')
#     # output_file_with_tickets = os.path.join(base_dir, 'alarms_having_tickets.xlsx')
#     # output_file_without_tickets = os.path.join(base_dir, 'alarms_not_having_tickets.xlsx')

#     df = load_excel_file(input_path)

#     if df is not None:
#         df = format_date_time_columns(df)
#         with_tickets, without_tickets = split_by_ticket_presence(df)

#         with_tickets = reset_serial_numbers(with_tickets)
#         without_tickets = reset_serial_numbers(without_tickets)

#         save_to_excel(with_tickets, output_file_with_tickets)
#         save_to_excel(without_tickets, output_file_without_tickets)

#         print("Files created successfully:")
#         print(f"- {output_file_with_tickets}")
#         print(f"- {output_file_without_tickets}")

# def create_custom_columns_excel():
#     input_path = './tickets_separation/alarms_not_having_tickets.xlsx'
#     output_file = './tickets_separation/alarms_not_having_tickets_custom_columns.xlsx'


#     # base_dir = os.path.dirname(__file__)  # gets the current alarms_processor.py path
#     # input_path = os.path.join(base_dir,  'alarms_not_having_tickets.xlsx')
#     # output_file = os.path.join(base_dir, 'alarms_not_having_tickets_custom_columns.xlsx')

#     df = load_excel_file(input_path)
#     if df is not None:
#         custom_df = pd.DataFrame({
#             'Date': df['Date'],
#             'No Of Days Trigerred': [''] * len(df),
#             'AlarmDescription': df['AlarmDescription'],
#             'Priority': df['Priority'],
#             'Count of AlarmDescription': [''] * len(df)
#         })
#         save_to_excel(custom_df, output_file)
#         print(f"Custom Excel file created successfully at: {output_file}")


# def add_duration_alarms_sheet():
#     custom_file = './tickets_separation/alarms_not_having_tickets_custom_columns.xlsx'
#     duration_file = './required_files/duration_alarms_list.xlsx'

#     # Load data
#     custom_df = pd.read_excel(custom_file, engine='openpyxl')
#     duration_df = pd.read_excel(duration_file, engine='openpyxl')

#     # Get list of duration alarms
#     duration_alarms_set = set(duration_df['AlarmDescription'].dropna().unique())

#     # Filter matching alarms
#     filtered_df = custom_df[custom_df['AlarmDescription'].isin(duration_alarms_set)].copy()

#     # Group and aggregate
#     grouped = filtered_df.groupby(['AlarmDescription', 'Priority'])
#     duration_data = []
#     for (alarm_desc, priority), group in grouped:
#         unique_dates = sorted(group['Date'].dropna().unique())
#         date_str = ' and '.join(unique_dates)
#         no_of_days = len(unique_dates)
#         count_of_alarm = len(group)
#         duration_data.append({
#             'Date': date_str,
#             'No Of Days Trigerred': no_of_days,
#             'AlarmDescription': alarm_desc,
#             'Priority': priority,
#             'Count of AlarmDescription': count_of_alarm
#         })

#     # Create new DataFrame and sort by 'Count of AlarmDescription' descending
#     duration_df_final = pd.DataFrame(duration_data, columns=[
#         'Date', 'No Of Days Trigerred', 'AlarmDescription', 'Priority', 'Count of AlarmDescription'
#     ])
#     duration_df_final.sort_values(by='Count of AlarmDescription', ascending=False, inplace=True)

#     # Write to new sheet
#     with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         duration_df_final.to_excel(writer, sheet_name='duration_alarms', index=False)

#     print("Sheet 'duration_alarms' added successfully.")



# def add_re_fine_tuning_alarms_sheet():
#     custom_file = './tickets_separation/alarms_not_having_tickets_custom_columns.xlsx'
#     re_fine_tuning_file = './required_files/re_fine_tuning_alarms_list.xlsx'

#     # Load data
#     custom_df = pd.read_excel(custom_file, engine='openpyxl')
#     re_fine_tuning_df = pd.read_excel(re_fine_tuning_file, engine='openpyxl')

#     # Get list of re-fine tuning alarms
#     re_fine_tuning_alarms_set = set(re_fine_tuning_df['AlarmDescription'].dropna().unique())

#     # Filter matching alarms
#     filtered_df = custom_df[custom_df['AlarmDescription'].isin(re_fine_tuning_alarms_set)].copy()

#     # Group and aggregate
#     grouped = filtered_df.groupby(['AlarmDescription', 'Priority'])
#     re_fine_tuning_data = []
#     for (alarm_desc, priority), group in grouped:
#         unique_dates = sorted(group['Date'].dropna().unique())
#         date_str = ' and '.join(unique_dates)
#         no_of_days = len(unique_dates)
#         count_of_alarm = len(group)
#         re_fine_tuning_data.append({
#             'Date': date_str,
#             'No Of Days Trigerred': no_of_days,
#             'AlarmDescription': alarm_desc,
#             'Priority': priority,
#             'Count of AlarmDescription': count_of_alarm
#         })

#     # Create new DataFrame and sort by Count of AlarmDescription descending
#     re_fine_tuning_df_final = pd.DataFrame(re_fine_tuning_data, columns=[
#         'Date', 'No Of Days Trigerred', 'AlarmDescription', 'Priority', 'Count of AlarmDescription'
#     ])
#     re_fine_tuning_df_final.sort_values(by='Count of AlarmDescription', ascending=False, inplace=True)

#     # Write to new sheet
#     with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         re_fine_tuning_df_final.to_excel(writer, sheet_name='re_fine_tuning_alarms', index=False)

#     print("Sheet 're_fine_tuning_alarms' added successfully.")



# def add_alarms_required_tickets_sheet():
#     custom_file = './tickets_separation/alarms_not_having_tickets_custom_columns.xlsx'
#     duration_file = './required_files/duration_alarms_list.xlsx'
#     re_fine_tuning_file = './required_files/re_fine_tuning_alarms_list.xlsx'

#     # Load data
#     custom_df = pd.read_excel(custom_file, engine='openpyxl')
#     duration_df = pd.read_excel(duration_file, engine='openpyxl')
#     re_fine_tuning_df = pd.read_excel(re_fine_tuning_file, engine='openpyxl')

#     # Get sets of alarm descriptions
#     duration_alarms_set = set(duration_df['AlarmDescription'].dropna().unique())
#     re_fine_tuning_alarms_set = set(re_fine_tuning_df['AlarmDescription'].dropna().unique())

#     # Exclude alarms already categorized
#     excluded_alarms = duration_alarms_set.union(re_fine_tuning_alarms_set)
#     remaining_df = custom_df[~custom_df['AlarmDescription'].isin(excluded_alarms)].copy()

#     # Group and aggregate
#     grouped = remaining_df.groupby(['AlarmDescription', 'Priority'])
#     required_tickets_data = []
#     for (alarm_desc, priority), group in grouped:
#         unique_dates = sorted(group['Date'].dropna().unique())
#         date_str = ' and '.join(unique_dates)
#         no_of_days = len(unique_dates)
#         count_of_alarm = len(group)
#         required_tickets_data.append({
#             'Date': date_str,
#             'No Of Days Trigerred': no_of_days,
#             'AlarmDescription': alarm_desc,
#             'Priority': priority,
#             'Count of AlarmDescription': count_of_alarm
#         })

#     # Create DataFrame and sort
#     required_tickets_df = pd.DataFrame(required_tickets_data, columns=[
#         'Date', 'No Of Days Trigerred', 'AlarmDescription', 'Priority', 'Count of AlarmDescription'
#     ])
#     required_tickets_df.sort_values(by='Count of AlarmDescription', ascending=False, inplace=True)

#     # Write to new sheet
#     with pd.ExcelWriter(custom_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         required_tickets_df.to_excel(writer, sheet_name='alarms_required_tickets', index=False)

#     print("Sheet 'alarms_required_tickets' added successfully.")






















