import getpass
import json
import os
from datetime import datetime, time as dtime, timezone, timedelta

import boto3
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


import pandas as pd
import os

OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "./alarm_report/reports")
AWS_REGION = os.environ.get("AWS_REGION", "us-east-1")

# Timezones the report can be generated in (chosen interactively at runtime).
UTC_TZ = timezone.utc
IST_TZ = timezone(timedelta(hours=5, minutes=30))

# Keywords used to guess severity from the alarm name when no tag is set.
# Order matters - first match wins, so put more specific words first.
SEVERITY_KEYWORDS = [
    ("critical", "Critical"),
    ("crit", "Critical"),
    ("high", "High"),
    ("medium", "Medium"),
    ("med", "Medium"),
    ("low", "Low"),
]

DEFAULT_SEVERITY = ""



def cleanup_alarm_reports():
    """
    Delete all existing Excel reports from the reports folder.
    """

    if not os.path.exists(OUTPUT_DIR):
        print(f"Reports folder not found: {OUTPUT_DIR}")
        return

    deleted_files = []

    for file_name in os.listdir(OUTPUT_DIR):

        if file_name.endswith(".xlsx"):

            file_path = os.path.join(
                OUTPUT_DIR,
                file_name
            )

            try:
                os.remove(file_path)
                deleted_files.append(file_name)

            except Exception as exc:
                print(
                    f"Failed to delete "
                    f"{file_name}: {exc}"
                )

    if deleted_files:

        print(
            "\nDeleted all old files:\n" +
            "\n".join(deleted_files)
        )

    else:
        print("No Excel files found to delete in reports.")



def choose_aws_profile():
    """
    Figure out which AWS SSO profile to use.

    Entering an invalid value will skip the profile.
    """

    env_profile = os.environ.get("AWS_PROFILE")

    if env_profile:
        print(f"Using AWS profile from AWS_PROFILE env var: {env_profile}")
        return env_profile

    available_profiles = boto3.Session().available_profiles

    if not available_profiles:
        return input(
            "No AWS profiles found locally. Enter the AWS profile name to use: "
        ).strip()

    print("Available AWS profiles:")

    for idx, profile in enumerate(available_profiles, start=1):
        print(f"  {idx}) {profile}")

    choice = input(
        "\nWhich profile do you want to log in with? "
        "(enter a number or the profile name): "
    ).strip()

    if choice.isdigit():

        profile_index = int(choice)

        if 1 <= profile_index <= len(available_profiles):
            return available_profiles[profile_index - 1]

    if choice in available_profiles:
        return choice

    print(
        f"\nSkipping profile selection. "
        f"'{choice}' is not a valid profile."
    )

    return None


def choose_timezone():
    """
    Ask the user which timezone the report's Date/Time columns (and the
    today/yesterday/week boundaries) should be based on: UTC or IST.
    """
    answer = input(
        "Which timezone do you want the report in? (UTC/IST): "
    ).strip().lower()

    if answer in ("ist", "india", "indian", "ist (utc+5:30)"):
        print("Using timezone: IST (UTC+5:30)")
        return IST_TZ, "IST"

    print("Using timezone: UTC")
    return UTC_TZ, "UTC"


def get_boto3_session():

    profile_name = choose_aws_profile()

    if profile_name is None:
        return None

    print(f"Using AWS SSO profile: {profile_name}")

    session = boto3.Session(
        profile_name=profile_name,
        region_name=AWS_REGION
    )

    try:

        identity = session.client("sts").get_caller_identity()

        print(
            f"Authenticated as: "
            f"{identity['Arn']}\n"
        )
    except Exception as exc:

        error_message = str(exc)

        if (
            "Token has expired" in error_message
            or "Error loading SSO Token" in error_message
        ):

            print(
                f"\nCould not authenticate using AWS SSO profile: "
                f"{error_message}"
            )

            print(
                f"\nFor example {profile_name} profile please run: "
                f"aws sso login --profile {profile_name}"
            )

            raise SystemExit(1)

        print(
            f"\nCould not authenticate using "
            f"AWS SSO profile: {error_message}"
        )

        return None
    
    return session



def get_account_id(sts_client):
    return sts_client.get_caller_identity()["Account"]

def get_account_label(session, account_id):
    """
    Human-readable label for this account, used in the sheet name and the
    output filename.

    Tries the account alias first (iam:ListAccountAliases). If no alias is
    set (or the caller doesn't have permission to read it), falls back to
    the raw account ID.
    """
    try:
        iam_client = session.client("iam")
        aliases = iam_client.list_account_aliases().get("AccountAliases", [])
        if aliases:
            return aliases[0]
    except Exception as exc:
        print(f"Could not read account alias, falling back to account ID: {exc}")

    return account_id


def sanitize_sheet_name(name):
    """Excel sheet names: max 31 chars, no []:*?/\\ characters."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    return name[:31]


def get_alarm_severity(cloudwatch_client, alarm_arn, alarm_name):
    """Severity priority: explicit tag > name keyword > default."""
    try:
        tags = cloudwatch_client.list_tags_for_resource(ResourceARN=alarm_arn).get("Tags", [])
        for tag in tags:
            if tag.get("Key", "").lower() in ("severity", "priority"):
                value = tag.get("Value", "").strip()
                if value:
                    return value.capitalize()
    except Exception as exc:
        print(f"Could not read tags for {alarm_name}: {exc}")

    lname = alarm_name.lower()
    for keyword, label in SEVERITY_KEYWORDS:
        if keyword in lname:
            return label

    return DEFAULT_SEVERITY


def get_all_alarms(cloudwatch_client):
    """
    Return every CloudWatch alarm (metric + composite) in the account/region,
    with its CURRENT state (OK / ALARM / INSUFFICIENT_DATA) and severity.
    Used to build the 'ok-inalarm' sheet, which lists all alarms regardless
    of whether they triggered today.
    """
    alarms = []
    paginator = cloudwatch_client.get_paginator("describe_alarms")
    for page in paginator.paginate():
        for alarm in page.get("MetricAlarms", []) + page.get("CompositeAlarms", []):
            alarm_name = alarm["AlarmName"]
            alarm_arn = alarm["AlarmArn"]
            current_state = alarm.get("StateValue", "UNKNOWN")

            alarms.append(
                {
                    "name": alarm_name,
                    "state": current_state,
                    "severity": get_alarm_severity(cloudwatch_client, alarm_arn, alarm_name),
                }
            )

    return alarms


def get_period_range(tz, period):
    """
    Supports:
        today
        yesterday
        week
        YYYY-MM-DD

    Examples:
        today
        yesterday
        week
        2026-08-14
    """

    now_tz = datetime.now(tz)
    today_date = now_tz.date()

    if period == "yesterday":

        target_date = today_date - timedelta(days=1)

        start = datetime.combine(
            target_date,
            dtime.min,
            tzinfo=tz
        )

        end = datetime.combine(
            target_date,
            dtime.max,
            tzinfo=tz
        )

    elif period == "week":

        start_date = today_date - timedelta(days=6)

        start = datetime.combine(
            start_date,
            dtime.min,
            tzinfo=tz
        )

        end = now_tz

    elif period == "today":

        start = datetime.combine(
            today_date,
            dtime.min,
            tzinfo=tz
        )

        end = now_tz

    else:
        # Assume specific date format YYYY-MM-DD

        try:

            target_date = datetime.strptime(
                period,
                "%Y-%m-%d"
            ).date()

            start = datetime.combine(
                target_date,
                dtime.min,
                tzinfo=tz
            )

            end = datetime.combine(
                target_date,
                dtime.max,
                tzinfo=tz
            )

        except ValueError:

            print(
                f"❌ Invalid date format: {period}\n"
                f"Expected format: YYYY-MM-DD"
            )

            return None, None

    return start, end


def get_alarm_events(cloudwatch_client, account_id, start, end, tz):
    """Return ALARM-state triggers between start and end, with Date/Time
    columns formatted in the given timezone."""
    events = []
    paginator = cloudwatch_client.get_paginator("describe_alarm_history")
    for page in paginator.paginate(
        HistoryItemType="StateUpdate",
        StartDate=start,
        EndDate=end,
        ScanBy="TimestampAscending",
    ):
        for item in page.get("AlarmHistoryItems", []):
            try:
                history_data = json.loads(item["HistoryData"])
            except (KeyError, json.JSONDecodeError):
                continue

            new_state = history_data.get("newState", {}).get("stateValue")
            if new_state != "ALARM":
                continue

            alarm_name = item["AlarmName"]
            alarm_arn = f"arn:aws:cloudwatch:{AWS_REGION}:{account_id}:alarm:{alarm_name}"
            timestamp = item["Timestamp"].astimezone(tz)  # convert to chosen timezone

            events.append(
                {
                    "date": timestamp.strftime("%Y-%m-%d"),
                    "time": timestamp.strftime("%H:%M:%S"),
                    "name": alarm_name,
                    "severity": get_alarm_severity(cloudwatch_client, alarm_arn, alarm_name),
                }
            )

    return events


def build_excel(events, all_alarms, account_label):
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    severity_colors = {
        "Critical": "FF0000",
        "High": "FFA500",
        "Medium": "FFFF00",
        "Low": "92D050",
    }
    state_colors = {
        "ALARM": "FF0000",
        "OK": "92D050",
        "INSUFFICIENT_DATA": "FFFF00",
    }

    # --- Sheet 1: Alarm Report (today's/period's ALARM triggers) ---
    ws = wb.active
    ws.title = sanitize_sheet_name(f"{account_label} Alarm Report")

    # headers = ["S.No", "Date", "Time", "AlarmDescription", "Priority"]
    headers = [
    "S.No",
    "Date",
    "Time",
    "AlarmDescription",
    "Priority",
    "HandledbyL2",
    "EscalatedtoL3/AppSupportteam",
    "Ticketnumber",
    "Status",
    ]
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    events.sort(key=lambda x: (x["date"], x["time"]))

    for idx, ev in enumerate(events, start=1):
        #row = [idx, ev["date"], ev["time"], ev["name"], ev["severity"]]
        row = [
            idx,
            ev["date"],
            ev["time"],
            ev["name"],
            ev["severity"],
            "",  # HandledbyL2
            "",  # EscalatedtoL3/AppSupportteam
            "",  # Ticketnumber
            "",  # Status
        ]
        ws.append(row)
        color = severity_colors.get(ev["severity"])
        if color:
            ws.cell(row=idx + 1, column=5).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    # for i, width in enumerate([8, 14, 12, 45, 12], start=1):
    #     ws.column_dimensions[get_column_letter(i)].width = width
    for i, width in enumerate([8, 14, 12, 45, 12, 18, 30, 18, 15],start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Sheet 2: ok-inalarm (every alarm, current state, regardless of
    #     whether it triggered in the report period) ---
    ws2 = wb.create_sheet(title=sanitize_sheet_name("ok-inalarm"))

    headers2 = ["S.No", "AlarmDescription", "CurrentState", "Priority"]
    ws2.append(headers2)
    for col_num in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    all_alarms_sorted = sorted(all_alarms, key=lambda x: x["name"])

    for idx, alarm in enumerate(all_alarms_sorted, start=1):
        row = [idx, alarm["name"], alarm["state"], alarm["severity"]]
        ws2.append(row)
        color = state_colors.get(alarm["state"])
        if color:
            ws2.cell(row=idx + 1, column=3).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
        sev_color = severity_colors.get(alarm["severity"])
        if sev_color:
            ws2.cell(row=idx + 1, column=4).fill = PatternFill(
                start_color=sev_color, end_color=sev_color, fill_type="solid"
            )

    for i, width in enumerate([8, 45, 16, 12], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    return wb


def alarm_reports():

    while True:

        session = get_boto3_session()

        if session is None:

            another = input(
                "\nDo you want to choose another AWS profile still? (y/n): "
            ).strip().lower()

            if another in ("y", "yes"):
                continue

            print("---Exiting alarm report generation---")
            break

        cloudwatch_client = session.client("cloudwatch")
        sts_client = session.client("sts")

        account_id = get_account_id(sts_client)
        account_label = get_account_label(session, account_id)

        tz, tz_label = choose_timezone()

        history_answer = input(
            "Enter one of the following:\n"
            "  today\n"
            "  yesterday\n"
            "  week\n"
            "  YYYY-MM-DD (example: 2026-08-14)\n"
            "Choice: "
        ).strip().lower()


        if history_answer in ("week", "1 week", "one week", "7", "7 days"):

            period = "week"
            period_label = "last 1 week"

        elif history_answer in ("yesterday", "yday", "y"):

            period = "yesterday"
            period_label = "yesterday"

        elif history_answer == "today":

            period = "today"
            period_label = "today"

        else:

            # User entered a specific date
            period = history_answer
            period_label = history_answer


        start, end = get_period_range(
            tz,
            period
        )
        if start is None or end is None:
            return

        events = get_alarm_events(
            cloudwatch_client,
            account_id,
            start,
            end,
            tz
        )

        count = len(events)

        print(
            f"Total alarm triggers for "
            f"{period_label} ({tz_label}): {count}"
        )

        print(
            "\nFetching all alarms for "
            "the ok-inalarm sheet..."
        )

        all_alarms = get_all_alarms(
            cloudwatch_client
        )

        print(
            f"Total alarms found in account: "
            f"{len(all_alarms)}"
        )

        if count == 0 and not all_alarms:

            print(
                f"No alarms for {period_label}, "
                f"and no alarms exist in the account. "
                f"No report generated."
            )

        else:

            wb = build_excel(
                events,
                all_alarms,
                account_label
            )

            os.makedirs(
                OUTPUT_DIR,
                exist_ok=True
            )

            date_str = datetime.now(tz).strftime(
                "%Y-%m-%d"
            )

            if period == "week":

                file_suffix = (
                    f"Last7Days_{date_str}"
                )

            elif period == "yesterday":

                yesterday_str = (
                    datetime.now(tz).date()
                    - timedelta(days=1)
                ).strftime("%Y-%m-%d")

                file_suffix = (
                    f"Yesterday_{yesterday_str}"
                )

            else:

                file_suffix = date_str

            safe_account_label = (
                account_label
                .replace(" ", "_")
                .replace("/", "-")
            )

            file_path = os.path.join(
                OUTPUT_DIR,
                f"AlarmReport_"
                f"{safe_account_label}_"
                f"{tz_label}_"
                f"{file_suffix}.xlsx"
            )

            wb.save(file_path)

            print(
                f"\nReport generated with "
                f"{count} alarm triggers and "
                f"{len(all_alarms)} total alarms -> "
                f"{os.path.abspath(file_path)}"
            )

        another = input(
            "\nDo you want to run another AWS profile? (y/n): "
        ).strip().lower()

        if another not in ("y", "yes"):

            print(
                "\nCompleted generating "
                "alarm reports."
            )

            break



def merge_alarm_reports():
    """
    Merge all AlarmReport_*.xlsx files into a single workbook.

    Output:
        All_AlarmReport.xlsx
    """

    if not os.path.exists(OUTPUT_DIR):
        print(f"Reports folder not found: {OUTPUT_DIR}")
        return

    report_files = sorted(
        [
            file_name
            for file_name in os.listdir(OUTPUT_DIR)
            if file_name.startswith("AlarmReport_")
            and file_name.endswith(".xlsx")
            and file_name != "All_AlarmReport.xlsx"
        ]
    )

    if not report_files:
        print("\nNo AlarmReport files found to merge....")
        return

    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "All Alarm Report"

    header_written = False
    serial_no = 1
    merged_files = []

    for file_name in report_files:

        file_path = os.path.join(
            OUTPUT_DIR,
            file_name
        )

        try:

            wb = openpyxl.load_workbook(file_path)

            source_sheet = None

            for sheet_name in wb.sheetnames:

                if sheet_name.endswith("Alarm Report"):
                    source_sheet = wb[sheet_name]
                    break

            if source_sheet is None:
                print(
                    f"Skipping {file_name} "
                    f"(Alarm Report sheet not found)"
                )
                continue

            rows = list(
                source_sheet.iter_rows(values_only=True)
            )

            if not rows:
                continue

            # Write header only once
            if not header_written:
                merged_ws.append(rows[0])
                header_written = True

            # Append data rows
            for row in rows[1:]:

                row = list(row)

                if row:
                    row[0] = serial_no

                merged_ws.append(row)

                serial_no += 1

            merged_files.append(file_name)

        except Exception as exc:
            print(
                f"Error processing "
                f"{file_name}: {exc}"
            )

    if not merged_files:
        print(
            "\nNo valid Alarm Report sheets were found. "
            "Merged workbook will not be created."
        )
        return

    print(
        "\nMerged: "
        + ", ".join(merged_files)
    )

    # Auto-size columns
    for column in merged_ws.columns:

        max_length = 0

        for cell in column:

            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        merged_ws.column_dimensions[
            get_column_letter(column[0].column)
        ].width = max_length + 5

    output_file = os.path.join(
        OUTPUT_DIR,
        "All_AlarmReport.xlsx"
    )

    # Delete old merged file if it exists
    if os.path.exists(output_file):
        os.remove(output_file)

        print(
            f"\nDeleted old file: "
            f"{os.path.abspath(output_file)}"
        )

    merged_wb.save(output_file)

    print(
        f"\nMerged Report Generated -> "
        f"{os.path.abspath(output_file)}"
    )


def priorities_and_handledbyl2():
    # File paths
    day_wise_file = "./alarm_report/reports/All_AlarmReport.xlsx"
    all_alarms_file = "./required_files/all_alarms_list.xlsx"

    output_file = "./alarm_report/reports/priorities_and_handledbyl2_filled_in_this_all_alarms_report.xlsx"

    # Delete old output file if it exists
    if os.path.exists(output_file):
        os.remove(output_file)
        print(f"Old file '{output_file}' deleted.")

    print("Processing priorities and HandledbyL2 columns...")

    # Check if input files exist
    if not os.path.exists(day_wise_file):
        print(f"❌ File not found: {day_wise_file}")
        print("Please generate All_AlarmReport.xlsx first to fill priorities and HandledbyL2 columns in priorities_and_handledbyl2_filled_in_this_all_alarms_report.xlsx.")
        return

    if not os.path.exists(all_alarms_file):
        print(f"❌ File not found: {all_alarms_file}")
        return

    # Load Excel files safely
    try:
        day_wise_df = pd.read_excel(day_wise_file)
        all_alarms_df = pd.read_excel(all_alarms_file)
    except Exception as exc:
        print(f"❌ Error reading Excel files: {exc}")
        return

    # Normalize column names
    day_wise_df.columns = day_wise_df.columns.str.strip()
    all_alarms_df.columns = all_alarms_df.columns.str.strip()

    # Replace empty strings with NaN
    day_wise_df[['Priority', 'HandledbyL2']] = day_wise_df[
        ['Priority', 'HandledbyL2']
    ].replace('', pd.NA)

    # Robust missing handling
    for col in ['Priority', 'HandledbyL2']:
        day_wise_df[col] = day_wise_df[col].astype('string').str.strip()
        day_wise_df[col] = day_wise_df[col].replace(
            {
                '': pd.NA,
                'nan': pd.NA,
                'None': pd.NA
            }
        )

    # Ensure target columns are string type
    day_wise_df['Priority'] = day_wise_df['Priority'].astype('string')
    day_wise_df['HandledbyL2'] = day_wise_df['HandledbyL2'].astype('string')

    # Preserve original AlarmDescription
    day_wise_df['OriginalAlarmDescription'] = day_wise_df['AlarmDescription']

    # Create normalized columns for matching
    day_wise_df['NormalizedAlarmDescription'] = (
        day_wise_df['AlarmDescription']
        .apply(lambda x: str(x).split('|')[0].strip().lower())
    )

    all_alarms_df['NormalizedAlarmDescription'] = (
        all_alarms_df['AlarmDescription']
        .apply(lambda x: str(x).split('|')[0].strip().lower())
    )

    # Create lookup dictionary
    alarm_priority_map = dict(
        zip(
            all_alarms_df['NormalizedAlarmDescription'],
            all_alarms_df['Priority']
        )
    )

    skipped_alarms = []

    # Iterate and update
    for idx, row in day_wise_df.iterrows():

        alarm_desc = row['NormalizedAlarmDescription']

        if pd.isna(row['Priority']) or pd.isna(row['HandledbyL2']):

            if alarm_desc in alarm_priority_map:

                if pd.isna(row['Priority']):
                    day_wise_df.at[idx, 'Priority'] = alarm_priority_map[alarm_desc]

                if pd.isna(row['HandledbyL2']):
                    day_wise_df.at[idx, 'HandledbyL2'] = "Yes"

            else:
                skipped_alarms.append(
                    row['OriginalAlarmDescription']
                )

        # Fallback
        if (
            pd.isna(day_wise_df.at[idx, 'HandledbyL2'])
            and
            pd.notna(day_wise_df.at[idx, 'Priority'])
        ):
            day_wise_df.at[idx, 'HandledbyL2'] = "Yes"

    # Format Date column
    if 'Date' in day_wise_df.columns:

        def format_date(x):
            if pd.isna(x) or str(x).strip() == '':
                return None

            try:
                dt = pd.to_datetime(x, errors='coerce')
                return (
                    dt.strftime('%m/%d/%Y')
                    if pd.notna(dt)
                    else str(x)
                )
            except Exception:
                return str(x)

        day_wise_df['Date'] = day_wise_df['Date'].apply(format_date)

    # Format Time column
    if 'Time' in day_wise_df.columns:

        def format_time(x):
            if pd.isna(x) or str(x).strip() == '':
                return None

            try:
                dt = pd.to_datetime(x, errors='coerce')
                return (
                    dt.strftime('%I:%M:%S %p')
                    if pd.notna(dt)
                    else str(x)
                )
            except Exception:
                return str(x)

        day_wise_df['Time'] = day_wise_df['Time'].apply(format_time)

    # Restore original AlarmDescription
    day_wise_df['AlarmDescription'] = day_wise_df['OriginalAlarmDescription']

    day_wise_df.drop(
        columns=[
            'NormalizedAlarmDescription',
            'OriginalAlarmDescription'
        ],
        inplace=True
    )

    # Save output file
    day_wise_df.to_excel(output_file, index=False)

    print(f"✅ Updated file saved as {output_file}")
    print(
        f"⚠ Skipped alarms (not found in all_alarms_list): "
        f"{len(skipped_alarms)}"
    )

    if skipped_alarms:
        print(skipped_alarms)