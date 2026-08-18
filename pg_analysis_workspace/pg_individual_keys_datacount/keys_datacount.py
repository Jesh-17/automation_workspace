
import sys
import re
from pathlib import Path
from typing import Tuple, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


def pg_individual_keys_datacount():
    """
    Build per-key, per-day count tables from a parsed logs Excel and
    save them into a new Excel file (one sheet per key) with table styling.

    Value fidelity:
      - Literal string "null"  -> stays "null" and IS counted
      - Empty/missing          -> empty cell (blank) and is NOT counted nor displayed
      - Other values           -> str(value), no trimming/normalization

    Keys:
      - Accept ANY column name exactly as in the file (incl. '[__type]').
      - Also tolerate user-typed variants (extra spaces, brackets added/removed, case differences, NBSP),
        and optional spaces around '|'.
      - A sheet is created for EVERY requested key:
          * If a column isn't found, a placeholder column is created so the sheet still appears.

    Date handling:
      - If 'Date' exists and is parseable -> per-day columns (YYYY-MM-DD ASC) + 'Grand Total'
      - If 'Date' missing/unparseable     -> only 'Grand Total' is produced

    Output:
      - 'pg_individual_keys_datacount.xlsx' (overwritten each run)
      - One sheet per requested key
      - Columns: date columns (if any) + 'Grand Total'
      - Rows: distinct labels with counts > 0, sorted by 'Grand Total' (desc),
              plus a bottom 'Grand Total' row.
    """

    # ----------------------------
    # Progress Bar
    # ----------------------------
    def print_progress(current: int, total: int, bar_width: int = 40) -> None:
        if total <= 0:
            return
        ratio = max(0.0, min(1.0, current / total))
        filled = int(bar_width * ratio)
        bar = '#' * filled + '-' * (bar_width - filled)
        percent = int(ratio * 100)
        sys.stdout.write(f"\rProcessing: [{bar}] {percent}%")
        sys.stdout.flush()
        if current >= total:
            sys.stdout.write("\n")

    # ----------------------------
    # Helpers (sheet name, input prompts)
    # ----------------------------
    def sanitize_sheet_name(name: str) -> str:
        # Replace invalids and also replace '|' defensively
        invalid = r'[:\\/\?\*\[\]\|]'
        clean = re.sub(invalid, "_", name or "Sheet1")
        return clean[:31] if clean else "Sheet1"

    def prompt_for_file_and_keys() -> Tuple[Path, List[str]]:
        file_in = input("📄 Enter Excel path (or press Enter to use 'sub_pg_refined.xlsx' in current folder): ").strip()
        if file_in:
            xlsx_path = Path(file_in)
        else:
            xlsx_path = Path("sub_pg_refined.xlsx")

        if not xlsx_path.exists():
            raise FileNotFoundError(
                f"Input file not found: {xlsx_path}\n"
                "Provide a valid path or place 'sub_pg_refined.xlsx' in the current directory."
            )

        print("🔑 Enter keys (separate with comma, NEWLINE, TAB, or 2+ spaces).")
        print("   Example (each is ONE column name):")
        print("   errors.0.httpStatusCode | errors.1.httpStatusCode    errors.0.code | errors.1.code")
        keys_raw = input("Keys: ").strip()
        if not keys_raw:
            raise ValueError("No keys provided. Please enter at least one key name.")

        # IMPORTANT:
        # - Preserve single spaces (because column names can contain ' | ').
        # - Do NOT split on single spaces.
        # - Split on commas, newlines, tabs, or 2+ spaces.
        parts = re.split(r"(?:,|\r?\n|\t| {2,})+", keys_raw)

        # Deduplicate while preserving order
        keys: List[str] = []
        for p in parts:
            k = p.strip()
            if k and k not in keys:
                keys.append(k)
        return xlsx_path, keys

    # ----------------------------
    # Data I/O helpers
    # ----------------------------
    def read_sheet(xlsx_path: Path) -> pd.DataFrame:
        """
        Preserve literal 'null' and empty cells:
          - keep_default_na=False => do NOT convert 'null' strings to NaN
          - na_filter=False       => do NOT auto-detect NA; blank cells stay empty strings
        Keep column names EXACTLY as in the sheet.
        """
        xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
        sheet = "TopKeys" if "TopKeys" in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(
            xls,
            sheet_name=sheet,
            engine="openpyxl",
            keep_default_na=False,
            na_filter=False,
            dtype=object  # keep mixed types; strings remain strings
        )
        df.columns = [str(c) for c in df.columns]
        return df

    def add_date_only_if_present(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        """
        If 'Date' exists, create 'DateOnly' (YYYY-MM-DD) and return sorted unique dates.
        If 'Date' missing/unparsable, return df unmodified and empty date list.
        """
        if "Date" not in df.columns:
            return df, []

        # Convert empty strings to NaN only for datetime parsing
        date_series = df["Date"].replace("", pd.NA)
        ts = pd.to_datetime(date_series, errors="coerce")

        if ts.notna().any():
            df = df.copy()
            df["DateOnly"] = ts.dt.strftime("%Y-%m-%d")  # NaT -> NaN
            vals = df["DateOnly"].dropna().astype(str)
            if not vals.empty:
                dt = pd.to_datetime(vals, format="%Y-%m-%d", errors="coerce").dropna()
                unique_sorted = pd.DatetimeIndex(pd.unique(dt)).sort_values()
                date_cols = unique_sorted.strftime("%Y-%m-%d").tolist()
                df["DateOnly"] = pd.Categorical(df["DateOnly"], categories=date_cols, ordered=True)
                return df, date_cols

        return df, []

    # ----------------------------
    # Label mapping (STRICT value fidelity)
    # ----------------------------
    def to_label(value) -> str:
        """
        - If value is a string, return it EXACTLY (so "null" stays "null"; "" stays "")
        - If value is missing (None/NaN), return "" (blank)
        - Else -> str(value) (no trimming/case change)
        """
        if isinstance(value, str):
            return value  # preserve exactly (incl. "null", "", spaces)
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        if value is None:
            return ""
        return str(value)

    def labels_series(raw_series: pd.Series) -> pd.Series:
        return raw_series.map(to_label)

    def first_seen_labels(raw_series: pd.Series) -> List[str]:
        order: List[str] = []
        for v in raw_series:
            lab = to_label(v)
            if lab not in order:
                order.append(lab)
        return order

    # ----------------------------
    # Counting utilities
    # ----------------------------
    def counts_table(df: pd.DataFrame, key: str, date_cols: List[str]) -> pd.DataFrame:
        """
        Compute counts per (label, DateOnly). If date_cols is empty, return empty here
        and let the caller build a Grand Total–only table.

        IMPORTANT: Exclude empty labels ("") from counts.
                   Literal "null" string is counted as "null".
        """
        if not date_cols or "DateOnly" not in df.columns:
            return pd.DataFrame()

        work = df[[key, "DateOnly"]].copy()
        work["__label__"] = labels_series(work[key])
        # Count only rows with valid DateOnly AND non-empty label
        work = work[(work["DateOnly"].notna()) & (work["__label__"] != "")]
        if work.empty:
            return pd.DataFrame()

        tab = (
            work.groupby(["__label__", "DateOnly"], observed=False)
                .size()
                .unstack("DateOnly", fill_value=0)
        )
        tab.index.name = key
        return tab

    def build_table_for_key(df: pd.DataFrame, key: str, date_cols: List[str]) -> pd.DataFrame:
        """
        Build the final table for a given key:

        - If date_cols exist: columns = date_cols + 'Grand Total'
        - If date_cols empty: columns = 'Grand Total' only
        - Rows include only labels with counts > 0 (rows with total 0 are dropped).
          * Empty labels "" are excluded from counts and are not shown.
          * Literal "null" (as text) is counted and shown if present.
        """
        if key not in df.columns:
            # Placeholder key -> return a minimal all-zero table
            if date_cols:
                total_series = {d: 0 for d in date_cols}
                total_series["Grand Total"] = 0
                out = pd.DataFrame([total_series], index=["Grand Total"])
                out.index.name = key
                return out
            else:
                base = pd.DataFrame({"Grand Total": [0]})
                base.index = ["Grand Total"]
                base.index.name = key
                return base

        # Preserve first-seen order for labels
        label_universe = first_seen_labels(df[key])

        # Try per-day counting first (excludes empty labels inside counts_table)
        tab = counts_table(df, key, date_cols)

        # -------- Case A: No date columns or no valid DateOnly -> Grand Total only ----------
        if tab.empty:
            # Count across all rows for non-empty labels only
            lbls_all = labels_series(df[key])
            lbls_no_empty = lbls_all[lbls_all != ""]              # EXCLUDE empty labels from counts
            totals = lbls_no_empty.value_counts(dropna=False)     # includes "null" if present

            # Keep only labels with count > 0, in first-seen order
            rows = []
            for lab in label_universe:
                cnt = int(totals.get(lab, 0))
                if cnt > 0:
                    rows.append((lab, cnt))

            # If nothing to show, still return a minimal table with just the bottom Grand Total row = 0
            if not rows:
                base = pd.DataFrame({"Grand Total": [0]})
                base.index = ["Grand Total"]
                base.index.name = key
                return base

            base = pd.DataFrame(rows, columns=[key, "Grand Total"]).set_index(key)
            base = base.sort_values(by="Grand Total", ascending=False)

            # Bottom Grand Total row
            total_row_val = int(base["Grand Total"].sum())
            base = pd.concat([base, pd.DataFrame({"Grand Total": [total_row_val]}, index=["Grand Total"])], axis=0)
            base.index.name = key
            return base

        # -------- Case B: Date columns exist ----------
        # Ensure date coverage + order
        for d in date_cols:
            if d not in tab.columns:
                tab[d] = 0
        tab = tab.reindex(columns=date_cols)

        # Include all labels (to preserve ordering), then fill with zeros
        tab = tab.reindex(label_universe).fillna(0).astype(int)

        # Add per-row Grand Total
        tab["Grand Total"] = tab.sum(axis=1)

        # Drop any rows that have Grand Total == 0
        tab = tab[tab["Grand Total"] > 0]

        # If everything got dropped, still return a minimal table with zeros
        if tab.empty:
            total_series = {d: 0 for d in date_cols}
            total_series["Grand Total"] = 0
            out = pd.DataFrame([total_series], index=["Grand Total"])
            out.index.name = key
            return out

        # Sort by Grand Total desc and append bottom Grand Total row
        tab = tab.sort_values(by="Grand Total", ascending=False)
        total_row = tab.sum(axis=0)
        total_row.name = "Grand Total"
        tab = pd.concat([tab, total_row.to_frame().T], axis=0)

        tab.index.name = key
        return tab

    def enforce_date_columns(table_df: pd.DataFrame, date_cols: List[str]) -> pd.DataFrame:
        if not date_cols:
            if "Grand Total" not in table_df.columns:
                table_df["Grand Total"] = table_df.sum(axis=1)
            return table_df[["Grand Total"]]
        for d in date_cols:
            if d not in table_df.columns:
                table_df[d] = 0
        expected = [*date_cols, "Grand Total"]
        if "Grand Total" not in table_df.columns:
            table_df["Grand Total"] = table_df.sum(axis=1)
        return table_df[expected]

    # ----------------------------
    # Key resolution (accept ANY column name; tolerate [__type], spaces, case, NBSP, spaces around '|')
    # ----------------------------
    def _strip_brackets_once(s: str) -> str:
        s = s.strip()
        if len(s) >= 2 and ((s[0] == '[' and s[-1] == ']') or (s[0] == '(' and s[-1] == ')') or (s[0] == '{' and s[-1] == '}')):
            return s[1:-1].strip()
        return s

    def _collapse_whitespace(s: str) -> str:
        # Normalize all whitespace (including NBSP) to a single space
        return re.sub(r"[\s\u00A0]+", " ", s).strip()

    def _normalize_pipes_ws(s: str) -> str:
        """
        Normalize spaces and NBSPs, then remove optional spaces around pipes:
        'a | b' -> 'a|b', 'a|b' -> 'a|b'
        """
        s = _collapse_whitespace(s)
        s = re.sub(r"\s*\|\s*", "|", s)
        return s

    def resolve_key_name(df_columns: List[str], requested: str) -> Optional[str]:
        """
        Resolve user-typed key to actual df column name.
        Matching order:
          1) exact, trimmed, bracket-stripped
          2) case-insensitive for (1)
          3) DF columns normalized by removing spaces around '|' and collapsing whitespace
          4) Same as (3) + one-layer bracket strip
          5) case-insensitive for (3) and (4)
        """
        cols = list(df_columns)

        req = requested
        req_trim = req.strip()
        req_nobr = _strip_brackets_once(req_trim)
        req_norm = _normalize_pipes_ws(req_trim)
        req_nobr_norm = _normalize_pipes_ws(req_nobr)

        # 1) direct exacts
        for cand in (req, req_trim, req_nobr):
            if cand in cols:
                return cand

        # 2) case-insensitive for direct
        lower_map = {c.lower(): c for c in cols}
        for cand in (req, req_trim, req_nobr):
            lc = cand.lower()
            if lc in lower_map:
                return lower_map[lc]

        # 3) normalize DF by pipes+ws
        df_norm_map = {_normalize_pipes_ws(c): c for c in cols}
        if req_norm in df_norm_map:
            return df_norm_map[req_norm]

        # 4) normalize DF by nobr + pipes+ws
        df_nobr_norm_map = {_normalize_pipes_ws(_strip_brackets_once(c)): c for c in cols}
        if req_nobr_norm in df_nobr_norm_map:
            return df_nobr_norm_map[req_nobr_norm]

        # 5) case-insensitive on normalized maps
        df_norm_lower = {k.lower(): v for k, v in df_norm_map.items()}
        df_nobr_norm_lower = {k.lower(): v for k, v in df_nobr_norm_map.items()}
        for lc in (req_norm.lower(), req_nobr_norm.lower()):
            if lc in df_norm_lower:
                return df_norm_lower[lc]
            if lc in df_nobr_norm_lower:
                return df_nobr_norm_lower[lc]

        return None

    # ----------------------------
    # Styling
    # ----------------------------
    def style_sheet_table(wb_path: Path, sheet_name: str) -> None:
        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            wb.save(wb_path)
            return

        ws = wb[sheet_name]
        last_row = ws.max_row
        last_col = ws.max_column

        ws.freeze_panes = "A2"

        if last_row < 2 or last_col < 1:
            wb.save(wb_path)
            return

        end_cell = f"{get_column_letter(last_col)}{last_row}"

        # Unique table name across workbook
        def _collect_existing_table_names(workbook):
            names = set()
            for w in workbook.worksheets:
                tbls = getattr(w, "tables", None)
                if isinstance(tbls, dict):
                    names.update(tbls.keys())
                elif tbls is not None:
                    for t in tbls:
                        if hasattr(t, "displayName"):
                            names.add(t.displayName)
                        elif isinstance(t, str):
                            names.add(t)
                for t in getattr(w, "_tables", []):
                    if hasattr(t, "displayName"):
                        names.add(t.displayName)
            return names

        existing_table_names = _collect_existing_table_names(wb)
        base_name = re.sub(r"\W+", "_", f"{sheet_name}_Table").strip("_") or "Table1"
        display_name = base_name[:60]
        suffix = 1
        while display_name in existing_table_names:
            suffix += 1
            display_name = (f"{base_name}_{suffix}")[:60]

        table = Table(displayName=display_name, ref=f"A1:{end_cell}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Wrap + vertical top alignment
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto width (scan up to 500 rows)
        for col_idx in range(1, last_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min(last_row, 500) + 1):
                cell_val = ws[f"{col_letter}{row_idx}"].value
                if cell_val is None:
                    continue
                max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)

        # Bold final row if 'Grand Total' row
        try:
            first_cell_val = ws[f"A{last_row}"].value
            if str(first_cell_val) == "Grand Total":
                for cell in ws[last_row]:
                    cell.font = Font(bold=True)
        except Exception:
            pass

        wb.save(wb_path)

    def fixed_output_name() -> Path:
        return Path("pg_individual_keys_datacount.xlsx")

    # ----------------------------
    # Execute
    # ----------------------------
    try:
        input_path, requested_keys = prompt_for_file_and_keys()
        print(f"📥 Input file: {input_path.resolve()}")

        # Load sheet (preserve 'null' strings and empty strings)
        df = read_sheet(input_path)

        # Column audit to reveal hidden characters
        print("🧭 Columns found in sheet (repr):")
        for c in df.columns:
            print("   -", repr(c))

        # DateOnly (if present)
        df, date_cols = add_date_only_if_present(df)

        if date_cols:
            print(f"📅 Included dates: {len(date_cols)} distinct days ({date_cols[0]} → {date_cols[-1]})")
        else:
            print("📅 No valid 'Date' column or no valid dates; will produce 'Grand Total' only.")

        print(f"🔑 Keys (requested): {', '.join(requested_keys)}")

        # Resolve keys, ALWAYS create a sheet
        df_cols = list(df.columns)
        resolved: List[Tuple[str, str]] = []  # (requested, actual-in-df)
        created_placeholders: List[str] = []

        for req_key in requested_keys:
            actual = resolve_key_name(df_cols, req_key)
            if actual is None:
                placeholder_col = req_key  # create placeholder so sheet is produced
                if placeholder_col not in df.columns:
                    df[placeholder_col] = pd.NA
                    df_cols.append(placeholder_col)
                    created_placeholders.append(req_key)
                resolved.append((req_key, placeholder_col))
            else:
                resolved.append((req_key, actual))

        print("🔎 Key resolution:")
        for req, act in resolved:
            note = " (placeholder)" if req in created_placeholders else ""
            print(f"   - {repr(req)} -> {repr(act)}{note}")

        out_path = fixed_output_name()

        total_steps = len(resolved) * 2  # write + style
        step = 0
        print_progress(step, total_steps)

        used_sheet_names = set()

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for (requested_name, actual_col) in resolved:
                table_df = build_table_for_key(df, actual_col, date_cols)
                table_df = enforce_date_columns(table_df, date_cols)

                base_name = sanitize_sheet_name(requested_name) or "Sheet1"
                sheet_name = base_name
                suffix = 1
                existing = used_sheet_names | set(writer.sheets.keys())
                if hasattr(writer, "book"):
                    existing |= {ws.title for ws in writer.book.worksheets}
                while sheet_name in existing:
                    suffix += 1
                    sheet_name = (base_name[:28] + f"_{suffix}")[:31]

                table_df.to_excel(writer, index=True, sheet_name=sheet_name, index_label=requested_name)
                used_sheet_names.add(sheet_name)

                step += 1
                print_progress(step, total_steps)

        # Style sheets
        for sheet_name in used_sheet_names:
            try:
                style_sheet_table(out_path, sheet_name)
            except Exception as se:
                print(f"\n⚠️ Styling skipped for sheet '{sheet_name}': {se}")
            step += 1
            print_progress(step, total_steps)

        print(f"\n✅ Done! Wrote analysis to '{out_path.name}'.")
        print(f"📄 Output path: {out_path.resolve()}")

        if created_placeholders:
            print("\nℹ️ Note: The following requested keys were not found as columns;")
            print("   placeholder sheets were created (all zeros) so you can see them anyway:")
            for k in created_placeholders:
                print("   -", k)

    except Exception as e:
        print("❌ Failed to run individual_keys_datacount.")
        print("Error:", str(e))
        import traceback
        traceback.print_exc()
        sys.exit(2)


# if __name__ == "__main__":
#     pg_individual_keys_datacount()








# import sys
# import re
# from pathlib import Path
# from typing import Tuple, List, Optional
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Alignment, Font


# def pg_individual_keys_datacount():
#     """
#     Build per-key, per-day count tables from a parsed logs Excel and
#     save them into a new Excel file (one sheet per key) with table styling.

#     Value fidelity:
#       - Literal string "null"  -> stays "null" and IS counted
#       - Empty/missing          -> empty cell (blank) and is NOT counted nor displayed
#       - Other values           -> str(value), no trimming/normalization

#     Keys:
#       - Accept ANY column name exactly as in the file (incl. '[__type]').
#       - Also tolerate user-typed variants (extra spaces, brackets added/removed, case differences, NBSP).
#       - A sheet is created for EVERY requested key:
#           * If a column isn't found, a placeholder column is created so the sheet still appears.

#     Date handling:
#       - If 'Date' exists and is parseable -> per-day columns (YYYY-MM-DD ASC) + 'Grand Total'
#       - If 'Date' missing/unparsable     -> only 'Grand Total' is produced

#     Output:
#       - 'pg_individual_keys_datacount.xlsx' (overwritten each run)
#       - One sheet per requested key
#       - Columns: date columns (if any) + 'Grand Total'
#       - Rows: distinct labels with counts > 0, sorted by 'Grand Total' (desc),
#               plus a bottom 'Grand Total' row.
#     """

#     # ----------------------------
#     # Progress Bar
#     # ----------------------------
#     def print_progress(current: int, total: int, bar_width: int = 40) -> None:
#         if total <= 0:
#             return
#         ratio = max(0.0, min(1.0, current / total))
#         filled = int(bar_width * ratio)
#         bar = '#' * filled + '-' * (bar_width - filled)
#         percent = int(ratio * 100)
#         sys.stdout.write(f"\rProcessing: [{bar}] {percent}%")
#         sys.stdout.flush()
#         if current >= total:
#             sys.stdout.write("\n")

#     # ----------------------------
#     # Helpers (sheet name, input prompts)
#     # ----------------------------
#     def sanitize_sheet_name(name: str) -> str:
#         # Replace invalids and also replace '|' defensively
#         invalid = r'[:\\/\?\*\[\]\|]'
#         clean = re.sub(invalid, "_", name or "Sheet1")
#         return clean[:31] if clean else "Sheet1"

#     def prompt_for_file_and_keys() -> Tuple[Path, List[str]]:
#         file_in = input("📄 Enter Excel path (or press Enter to use 'sub_pg_refined.xlsx' in current folder): ").strip()
#         if file_in:
#             xlsx_path = Path(file_in)
#         else:
#             xlsx_path = Path("sub_pg_refined.xlsx")

#         if not xlsx_path.exists():
#             raise FileNotFoundError(
#                 f"Input file not found: {xlsx_path}\n"
#                 "Provide a valid path or place 'sub_pg_refined.xlsx' in the current directory."
#             )

#         print("🔑 Enter keys (separate with comma, NEWLINE, TAB, or 2+ spaces).")
#         print("   Example (each is ONE column name):")
#         print("   errors.0.httpStatusCode | errors.1.httpStatusCode    errors.0.code | errors.1.code")
#         keys_raw = input("Keys: ").strip()
#         if not keys_raw:
#             raise ValueError("No keys provided. Please enter at least one key name.")

#         # Preserve single spaces inside column names (esp. around '|').
#         # Split on commas, newlines, tabs, or 2+ spaces.
#         parts = re.split(r"(?:,|\r?\n|\t| {2,})+", keys_raw)

#         # Deduplicate while preserving order
#         keys: List[str] = []
#         for p in parts:
#             k = p.strip()
#             if k and k not in keys:
#                 keys.append(k)
#         return xlsx_path, keys

#     # ----------------------------
#     # Data I/O helpers
#     # ----------------------------
#     def read_sheet(xlsx_path: Path) -> pd.DataFrame:
#         """
#         Preserve literal 'null' and empty cells:
#           - keep_default_na=False => do NOT convert 'null' strings to NaN
#           - na_filter=False       => do NOT auto-detect NA; blank cells stay empty strings
#         Keep column names EXACTLY as in the sheet.
#         """
#         xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
#         sheet = "TopKeys" if "TopKeys" in xls.sheet_names else xls.sheet_names[0]
#         df = pd.read_excel(
#             xls,
#             sheet_name=sheet,
#             engine="openpyxl",
#             keep_default_na=False,
#             na_filter=False,
#             dtype=object  # keep mixed types; strings remain strings
#         )
#         df.columns = [str(c) for c in df.columns]
#         return df

#     def add_date_only_if_present(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
#         """
#         If 'Date' exists, create 'DateOnly' (YYYY-MM-DD) and return sorted unique dates.
#         If 'Date' missing/unparsable, return df unmodified and empty date list.
#         """
#         if "Date" not in df.columns:
#             return df, []

#         # Convert empty strings to NaN only for datetime parsing
#         date_series = df["Date"].replace("", pd.NA)
#         ts = pd.to_datetime(date_series, errors="coerce")

#         if ts.notna().any():
#             df = df.copy()
#             df["DateOnly"] = ts.dt.strftime("%Y-%m-%d")  # NaT -> NaN
#             vals = df["DateOnly"].dropna().astype(str)
#             if not vals.empty:
#                 dt = pd.to_datetime(vals, format="%Y-%m-%d", errors="coerce").dropna()
#                 unique_sorted = pd.DatetimeIndex(pd.unique(dt)).sort_values()
#                 date_cols = unique_sorted.strftime("%Y-%m-%d").tolist()
#                 df["DateOnly"] = pd.Categorical(df["DateOnly"], categories=date_cols, ordered=True)
#                 return df, date_cols

#         return df, []

#     # ----------------------------
#     # Label mapping (STRICT value fidelity)
#     # ----------------------------
#     def to_label(value) -> str:
#         """
#         - If value is a string, return it EXACTLY (so "null" stays "null"; "" stays "")
#         - If value is missing (None/NaN), return "" (blank)
#         - Else -> str(value) (no trimming/case change)
#         """
#         if isinstance(value, str):
#             return value  # preserve exactly (incl. "null", "", spaces)
#         try:
#             if pd.isna(value):
#                 return ""
#         except Exception:
#             pass
#         if value is None:
#             return ""
#         return str(value)

#     def labels_series(raw_series: pd.Series) -> pd.Series:
#         return raw_series.map(to_label)

#     def first_seen_labels(raw_series: pd.Series) -> List[str]:
#         order: List[str] = []
#         for v in raw_series:
#             lab = to_label(v)
#             if lab not in order:
#                 order.append(lab)
#         return order

#     # ----------------------------
#     # Counting utilities
#     # ----------------------------
#     def counts_table(df: pd.DataFrame, key: str, date_cols: List[str]) -> pd.DataFrame:
#         """
#         Compute counts per (label, DateOnly). If date_cols is empty, return empty here
#         and let the caller build a Grand Total–only table.

#         IMPORTANT: Exclude empty labels ("") from counts.
#                    Literal "null" string is counted as "null".
#         """
#         if not date_cols or "DateOnly" not in df.columns:
#             return pd.DataFrame()

#         work = df[[key, "DateOnly"]].copy()
#         work["__label__"] = labels_series(work[key])
#         # Count only rows with valid DateOnly AND non-empty label
#         work = work[(work["DateOnly"].notna()) & (work["__label__"] != "")]
#         if work.empty:
#             return pd.DataFrame()

#         tab = (
#             work.groupby(["__label__", "DateOnly"], observed=False)
#                 .size()
#                 .unstack("DateOnly", fill_value=0)
#         )
#         tab.index.name = key
#         return tab

#     def build_table_for_key(df: pd.DataFrame, key: str, date_cols: List[str]) -> pd.DataFrame:
#         """
#         Build the final table for a given key:

#         - If date_cols exist: columns = date_cols + 'Grand Total'
#         - If date_cols empty: columns = 'Grand Total' only
#         - Rows include only labels with counts > 0 (rows with total 0 are dropped).
#           * Empty labels "" are excluded from counts and are not shown.
#           * Literal "null" (as text) is counted and shown if present.
#         """
#         if key not in df.columns:
#             # Placeholder key -> return a minimal all-zero table
#             if date_cols:
#                 total_series = {d: 0 for d in date_cols}
#                 total_series["Grand Total"] = 0
#                 out = pd.DataFrame([total_series], index=["Grand Total"])
#                 out.index.name = key
#                 return out
#             else:
#                 base = pd.DataFrame({"Grand Total": [0]})
#                 base.index = ["Grand Total"]
#                 base.index.name = key
#                 return base

#         # Preserve first-seen order for labels
#         label_universe = first_seen_labels(df[key])

#         # Try per-day counting first (excludes empty labels inside counts_table)
#         tab = counts_table(df, key, date_cols)

#         # -------- Case A: No date columns or no valid DateOnly -> Grand Total only ----------
#         if tab.empty:
#             # Count across all rows for non-empty labels only
#             lbls_all = labels_series(df[key])
#             lbls_no_empty = lbls_all[lbls_all != ""]              # EXCLUDE empty labels from counts
#             totals = lbls_no_empty.value_counts(dropna=False)     # includes "null" if present

#             # Keep only labels with count > 0, in first-seen order
#             rows = []
#             for lab in label_universe:
#                 cnt = int(totals.get(lab, 0))
#                 if cnt > 0:
#                     rows.append((lab, cnt))

#             # If nothing to show, still return a minimal table with just the bottom Grand Total row = 0
#             if not rows:
#                 base = pd.DataFrame({"Grand Total": [0]})
#                 base.index = ["Grand Total"]
#                 base.index.name = key
#                 return base

#             base = pd.DataFrame(rows, columns=[key, "Grand Total"]).set_index(key)
#             base = base.sort_values(by="Grand Total", ascending=False)

#             # Bottom Grand Total row
#             total_row_val = int(base["Grand Total"].sum())
#             base = pd.concat([base, pd.DataFrame({"Grand Total": [total_row_val]}, index=["Grand Total"])], axis=0)
#             base.index.name = key
#             return base

#         # -------- Case B: Date columns exist ----------
#         # Ensure date coverage + order
#         for d in date_cols:
#             if d not in tab.columns:
#                 tab[d] = 0
#         tab = tab.reindex(columns=date_cols)

#         # Include all labels (to preserve ordering), then fill with zeros
#         tab = tab.reindex(label_universe).fillna(0).astype(int)

#         # Add per-row Grand Total
#         tab["Grand Total"] = tab.sum(axis=1)

#         # Drop any rows that have Grand Total == 0
#         tab = tab[tab["Grand Total"] > 0]

#         # If everything got dropped, still return a minimal table with zeros
#         if tab.empty:
#             total_series = {d: 0 for d in date_cols}
#             total_series["Grand Total"] = 0
#             out = pd.DataFrame([total_series], index=["Grand Total"])
#             out.index.name = key
#             return out

#         # Sort by Grand Total desc and append bottom Grand Total row
#         tab = tab.sort_values(by="Grand Total", ascending=False)
#         total_row = tab.sum(axis=0)
#         total_row.name = "Grand Total"
#         tab = pd.concat([tab, total_row.to_frame().T], axis=0)

#         tab.index.name = key
#         return tab

#     def enforce_date_columns(table_df: pd.DataFrame, date_cols: List[str]) -> pd.DataFrame:
#         if not date_cols:
#             if "Grand Total" not in table_df.columns:
#                 table_df["Grand Total"] = table_df.sum(axis=1)
#             return table_df[["Grand Total"]]
#         for d in date_cols:
#             if d not in table_df.columns:
#                 table_df[d] = 0
#         expected = [*date_cols, "Grand Total"]
#         if "Grand Total" not in table_df.columns:
#             table_df["Grand Total"] = table_df.sum(axis=1)
#         return table_df[expected]

#     # ----------------------------
#     # Key resolution (accept ANY column name; tolerate [__type], spaces, case, NBSP, spaces around '|')
#     # ----------------------------
#     def _strip_brackets_once(s: str) -> str:
#         s = s.strip()
#         if len(s) >= 2 and ((s[0] == '[' and s[-1] == ']') or (s[0] == '(' and s[-1] == ')') or (s[0] == '{' and s[-1] == '}')):
#             return s[1:-1].strip()
#         return s

#     def _collapse_whitespace(s: str) -> str:
#         # Normalize all whitespace (including NBSP) to a single space
#         return re.sub(r"[\s\u00A0]+", " ", s).strip()

#     def _normalize_pipes_ws(s: str) -> str:
#         """
#         Normalize spaces and NBSPs, then remove optional spaces around pipes:
#         'a | b' -> 'a|b', 'a|b' -> 'a|b'
#         """
#         s = _collapse_whitespace(s)
#         s = re.sub(r"\s*\|\s*", "|", s)
#         return s

#     def resolve_key_name(df_columns: List[str], requested: str) -> Optional[str]:
#         """
#         Resolve user-typed key to actual df column name.
#         Matching order:
#           1) exact, trimmed, bracket-stripped
#           2) case-insensitive for (1)
#           3) DF columns normalized by removing spaces around '|' and collapsing whitespace
#           4) Same as (3) + one-layer bracket strip
#           5) case-insensitive for (3) and (4)
#         """
#         cols = list(df_columns)

#         req = requested
#         req_trim = req.strip()
#         req_nobr = _strip_brackets_once(req_trim)
#         req_norm = _normalize_pipes_ws(req_trim)
#         req_nobr_norm = _normalize_pipes_ws(req_nobr)

#         # 1) direct exacts
#         for cand in (req, req_trim, req_nobr):
#             if cand in cols:
#                 return cand

#         # 2) case-insensitive for direct
#         lower_map = {c.lower(): c for c in cols}
#         for cand in (req, req_trim, req_nobr):
#             lc = cand.lower()
#             if lc in lower_map:
#                 return lower_map[lc]

#         # 3) normalize DF by pipes+ws
#         df_norm_map = {_normalize_pipes_ws(c): c for c in cols}
#         if req_norm in df_norm_map:
#             return df_norm_map[req_norm]

#         # 4) normalize DF by nobr + pipes+ws
#         df_nobr_norm_map = {_normalize_pipes_ws(_strip_brackets_once(c)): c for c in cols}
#         if req_nobr_norm in df_nobr_norm_map:
#             return df_nobr_norm_map[req_nobr_norm]

#         # 5) case-insensitive on normalized maps
#         df_norm_lower = {k.lower(): v for k, v in df_norm_map.items()}
#         df_nobr_norm_lower = {k.lower(): v for k, v in df_nobr_norm_map.items()}
#         for lc in (req_norm.lower(), req_nobr_norm.lower()):
#             if lc in df_norm_lower:
#                 return df_norm_lower[lc]
#             if lc in df_nobr_norm_lower:
#                 return df_nobr_norm_lower[lc]

#         return None

#     # ----------------------------
#     # Styling
#     # ----------------------------
#     def style_sheet_table(wb_path: Path, sheet_name: str) -> None:
#         wb = load_workbook(wb_path)
#         if sheet_name not in wb.sheetnames:
#             wb.save(wb_path)
#             return

#         ws = wb[sheet_name]
#         last_row = ws.max_row
#         last_col = ws.max_column

#         ws.freeze_panes = "A2"

#         if last_row < 2 or last_col < 1:
#             wb.save(wb_path)
#             return

#         end_cell = f"{get_column_letter(last_col)}{last_row}"

#         # Unique table name across workbook
#         def _collect_existing_table_names(workbook):
#             names = set()
#             for w in workbook.worksheets:
#                 tbls = getattr(w, "tables", None)
#                 if isinstance(tbls, dict):
#                     names.update(tbls.keys())
#                 elif tbls is not None:
#                     for t in tbls:
#                         if hasattr(t, "displayName"):
#                             names.add(t.displayName)
#                         elif isinstance(t, str):
#                             names.add(t)
#                 for t in getattr(w, "_tables", []):
#                     if hasattr(t, "displayName"):
#                         names.add(t.displayName)
#             return names

#         existing_table_names = _collect_existing_table_names(wb)
#         base_name = re.sub(r"\W+", "_", f"{sheet_name}_Table").strip("_") or "Table1"
#         display_name = base_name[:60]
#         suffix = 1
#         while display_name in existing_table_names:
#             suffix += 1
#             display_name = (f"{base_name}_{suffix}")[:60]

#         table = Table(displayName=display_name, ref=f"A1:{end_cell}")
#         style = TableStyleInfo(
#             name="TableStyleMedium9",
#             showFirstColumn=False,
#             showLastColumn=False,
#             showRowStripes=True,
#             showColumnStripes=False,
#         )
#         table.tableStyleInfo = style
#         ws.add_table(table)

#         # Wrap + vertical top alignment
#         for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
#             for cell in row:
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#         # Auto width (scan up to 500 rows)
#         for col_idx in range(1, last_col + 1):
#             col_letter = get_column_letter(col_idx)
#             max_len = 0
#             for row_idx in range(1, min(last_row, 500) + 1):
#                 cell_val = ws[f"{col_letter}{row_idx}"].value
#                 if cell_val is None:
#                     continue
#                 max_len = max(max_len, len(str(cell_val)))
#             ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)

#         # Bold final row if 'Grand Total' row
#         try:
#             first_cell_val = ws[f"A{last_row}"].value
#             if str(first_cell_val) == "Grand Total":
#                 for cell in ws[last_row]:
#                     cell.font = Font(bold=True)
#         except Exception:
#             pass

#         wb.save(wb_path)

#     def fixed_output_name() -> Path:
#         return Path("pg_individual_keys_datacount.xlsx")

#     # ----------------------------
#     # Execute
#     # ----------------------------
#     try:
#         input_path, requested_keys = prompt_for_file_and_keys()
#         print(f"📥 Input file: {input_path.resolve()}")

#         # Load sheet (preserve 'null' strings and empty strings)
#         df = read_sheet(input_path)

#         # Column audit to reveal hidden characters
#         print("🧭 Columns found in sheet (repr):")
#         for c in df.columns:
#             print("   -", repr(c))

#         # DateOnly (if present)
#         df, date_cols = add_date_only_if_present(df)

#         if date_cols:
#             print(f"📅 Included dates: {len(date_cols)} distinct days ({date_cols[0]} → {date_cols[-1]})")
#         else:
#             print("📅 No valid 'Date' column or no valid dates; will produce 'Grand Total' only.")

#         print(f"🔑 Keys (requested): {', '.join(requested_keys)}")

#         # Resolve keys, ALWAYS create a sheet
#         df_cols = list(df.columns)
#         from typing import Tuple as _Tuple
#         resolved: List[_Tuple[str, str]] = []  # (requested, actual-in-df)
#         created_placeholders: List[str] = []

#         for req_key in requested_keys:
#             actual = resolve_key_name(df_cols, req_key)
#             if actual is None:
#                 placeholder_col = req_key  # create placeholder so sheet is produced
#                 if placeholder_col not in df.columns:
#                     df[placeholder_col] = pd.NA
#                     df_cols.append(placeholder_col)
#                     created_placeholders.append(req_key)
#                 resolved.append((req_key, placeholder_col))
#             else:
#                 resolved.append((req_key, actual))

#         print("🔎 Key resolution:")
#         for req, act in resolved:
#             note = " (placeholder)" if req in created_placeholders else ""
#             print(f"   - {repr(req)} -> {repr(act)}{note}")

#         out_path = fixed_output_name()

#         total_steps = len(resolved) * 2  # write + style
#         step = 0
#         print_progress(step, total_steps)

#         used_sheet_names = set()

#         with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
#             for (requested_name, actual_col) in resolved:
#                 table_df = build_table_for_key(df, actual_col, date_cols)
#                 table_df = enforce_date_columns(table_df, date_cols)

#                 base_name = sanitize_sheet_name(requested_name) or "Sheet1"
#                 sheet_name = base_name
#                 suffix = 1
#                 existing = used_sheet_names | set(writer.sheets.keys())
#                 if hasattr(writer, "book"):
#                     existing |= {ws.title for ws in writer.book.worksheets}
#                 while sheet_name in existing:
#                     suffix += 1
#                     sheet_name = (base_name[:28] + f"_{suffix}")[:31]

#                 table_df.to_excel(writer, index=True, sheet_name=sheet_name, index_label=requested_name)
#                 used_sheet_names.add(sheet_name)

#                 step += 1
#                 print_progress(step, total_steps)

#         # Style sheets
#         for sheet_name in used_sheet_names:
#             try:
#                 style_sheet_table(out_path, sheet_name)
#             except Exception as se:
#                 print(f"\n⚠️ Styling skipped for sheet '{sheet_name}': {se}")
#             step += 1
#             print_progress(step, total_steps)

#         print(f"\n✅ Done! Wrote analysis to '{out_path.name}'.")
#         print(f"📄 Output path: {out_path.resolve()}")

#         if created_placeholders:
#             print("\nℹ️ Note: The following requested keys were not found as columns;")
#             print("   placeholder sheets were created (all zeros) so you can see them anyway:")
#             for k in created_placeholders:
#                 print("   -", k)

#     except Exception as e:
#         print("❌ Failed to run individual_keys_datacount.")
#         print("Error:", str(e))
#         import traceback
#         traceback.print_exc()
#         sys.exit(2)




