
import sys
import re
from pathlib import Path
from typing import Tuple, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


def sub_dar_individual_keys_datacount():
    """
    Build per-key, per-day count tables from a parsed logs Excel and
    save them into a new Excel file (one sheet per key) with table styling.

    Value fidelity:
      - Literal string "null"  -> stays "null" and IS counted
      - Empty/missing          -> empty cell (blank) and is NOT counted nor displayed
      - Other values           -> str(value), no trimming/normalization

    Keys:
      - Accept ANY column name exactly as in the file (incl. '[__type]').
      - Also tolerate user-typed variants (extra spaces, brackets added/removed, case differences, NBSP).
      - A sheet is created for EVERY requested key:
          * If a column isn't found, a placeholder column is created so the sheet still appears.

    Date handling:
      - If 'Date' exists and is parseable -> per-day columns (YYYY-MM-DD ASC) + 'Grand Total'
      - If 'Date' missing/unparseable     -> only 'Grand Total' is produced

    Output:
      - 'sub_dar_individual_keys_datacount.xlsx' (overwritten each run)
      - One sheet per requested key
      - Columns: date columns (if any) + 'Grand Total'
      - Rows: distinct labels with counts > 0, sorted by 'Grand Total' (desc),
              plus a bottom 'Grand Total' row.
    """

    # ----------------------------
    # Progress Bar
    # ----------------------------
    def print_progress(current: int, total: int, bar_width: int = 40) -> None:
        if total <= 0:
            return
        ratio = max(0.0, min(1.0, current / total))
        filled = int(bar_width * ratio)
        bar = '#' * filled + '-' * (bar_width - filled)
        percent = int(ratio * 100)
        sys.stdout.write(f"\rProcessing: [{bar}] {percent}%")
        sys.stdout.flush()
        if current >= total:
            sys.stdout.write("\n")

    # ----------------------------
    # Helpers (sheet name, input prompts)
    # ----------------------------
    def sanitize_sheet_name(name: str) -> str:
        invalid = r'[:\\/\?\*\[\]]'
        clean = re.sub(invalid, "_", name or "Sheet1")
        return clean[:31] if clean else "Sheet1"

    def prompt_for_file_and_keys() -> Tuple[Path, List[str]]:
        file_in = input("📄 Enter Excel path (or press Enter to use 'sub_dar_refined.xlsx' in current folder): ").strip()
        if file_in:
            xlsx_path = Path(file_in)
        else:
            xlsx_path = Path("sub_dar_refined.xlsx")

        if not xlsx_path.exists():
            raise FileNotFoundError(
                f"Input file not found: {xlsx_path}\n"
                "Provide a valid path or place 'sub_dar_refined.xlsx' in the current directory."
            )

        keys_raw = input("🔑 Enter keys (comma or space separated): ").strip()
        if not keys_raw:
            raise ValueError("No keys provided. Please enter at least one key name.")
        parts = re.split(r"[,\s]+", keys_raw)  # split on spaces/newlines/commas
        keys: List[str] = []
        for p in parts:
            if p:
                k = p.strip()
                if k and k not in keys:
                    keys.append(k)
        return xlsx_path, keys

    # ----------------------------
    # Data I/O helpers
    # ----------------------------
    def read_sheet(xlsx_path: Path) -> pd.DataFrame:
        """
        Preserve literal 'null' and empty cells:
          - keep_default_na=False => do NOT convert 'null' strings to NaN
          - na_filter=False       => do NOT auto-detect NA; blank cells stay empty strings
        Keep column names EXACTLY as in the sheet.
        """
        xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
        sheet = "TopKeys" if "TopKeys" in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(
            xls,
            sheet_name=sheet,
            engine="openpyxl",
            keep_default_na=False,
            na_filter=False,
            dtype=object  # keep mixed types; strings remain strings
        )
        df.columns = [str(c) for c in df.columns]
        return df

    def add_date_only_if_present(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        """
        If 'Date' exists, create 'DateOnly' (YYYY-MM-DD) and return sorted unique dates.
        If 'Date' missing/unparsable, return df unmodified and empty date list.
        """
        if "Date" not in df.columns:
            return df, []

        # Convert empty strings to NaN only for datetime parsing
        date_series = df["Date"].replace("", pd.NA)
        ts = pd.to_datetime(date_series, errors="coerce")

        if ts.notna().any():
            df = df.copy()
            df["DateOnly"] = ts.dt.strftime("%Y-%m-%d")  # NaT -> NaN
            vals = df["DateOnly"].dropna().astype(str)
            if not vals.empty:
                dt = pd.to_datetime(vals, format="%Y-%m-%d", errors="coerce").dropna()
                unique_sorted = pd.DatetimeIndex(pd.unique(dt)).sort_values()
                date_cols = unique_sorted.strftime("%Y-%m-%d").tolist()
                df["DateOnly"] = pd.Categorical(df["DateOnly"], categories=date_cols, ordered=True)
                return df, date_cols

        return df, []

    # ----------------------------
    # Label mapping (STRICT value fidelity)
    # ----------------------------
    def to_label(value) -> str:
        """
        - If value is a string, return it EXACTLY (so "null" stays "null"; "" stays "")
        - If value is missing (None/NaN), return "" (blank)
        - Else -> str(value) (no trimming/case change)
        """
        if isinstance(value, str):
            return value  # preserve exactly (incl. "null", "", spaces)
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        if value is None:
            return ""
        return str(value)

    def labels_series(raw_series: pd.Series) -> pd.Series:
        return raw_series.map(to_label)

    def first_seen_labels(raw_series: pd.Series) -> List[str]:
        order: List[str] = []
        for v in raw_series:
            lab = to_label(v)
            if lab not in order:
                order.append(lab)
        return order

    # ----------------------------
    # Counting utilities
    # ----------------------------
    def counts_table(df: pd.DataFrame, key: str, date_cols: List[str]) -> pd.DataFrame:
        """
        Compute counts per (label, DateOnly). If date_cols is empty, return empty here
        and let the caller build a Grand Total–only table.

        IMPORTANT: Exclude empty labels ("") from counts.
                   Literal "null" string is counted as "null".
        """
        if not date_cols or "DateOnly" not in df.columns:
            return pd.DataFrame()

        work = df[[key, "DateOnly"]].copy()
        work["__label__"] = labels_series(work[key])
        # Count only rows with valid DateOnly AND non-empty label
        work = work[(work["DateOnly"].notna()) & (work["__label__"] != "")]
        if work.empty:
            return pd.DataFrame()

        tab = (
            work.groupby(["__label__", "DateOnly"], observed=False)
                .size()
                .unstack("DateOnly", fill_value=0)
        )
        tab.index.name = key
        return tab

    def build_table_for_key(df: pd.DataFrame, key: str, date_cols: List[str]) -> pd.DataFrame:
        """
        Build the final table for a given key:

        - If date_cols exist: columns = date_cols + 'Grand Total'
        - If date_cols empty: columns = 'Grand Total' only
        - Rows include only labels with counts > 0 (rows with total 0 are dropped).
          * Empty labels "" are excluded from counts and are not shown.
          * Literal "null" (as text) is counted and shown if present.
        """
        if key not in df.columns:
            raise KeyError(f"Key '{key}' not found in sheet columns.")

        # Preserve first-seen order for labels
        label_universe = first_seen_labels(df[key])

        # Try per-day counting first (excludes empty labels inside counts_table)
        tab = counts_table(df, key, date_cols)

        # -------- Case A: No date columns or no valid DateOnly -> Grand Total only ----------
        if tab.empty:
            # Count across all rows for non-empty labels only
            lbls_all = labels_series(df[key])
            lbls_no_empty = lbls_all[lbls_all != ""]              # EXCLUDE empty labels from counts
            totals = lbls_no_empty.value_counts(dropna=False)     # includes "null" if present

            # Keep only labels with count > 0, in first-seen order
            rows = []
            for lab in label_universe:
                cnt = int(totals.get(lab, 0))
                if cnt > 0:
                    rows.append((lab, cnt))

            # If nothing to show, still return a minimal table with just the bottom Grand Total row = 0
            if not rows:
                base = pd.DataFrame({"Grand Total": [0]})
                base.index = ["Grand Total"]
                base.index.name = key
                return base

            base = pd.DataFrame(rows, columns=[key, "Grand Total"]).set_index(key)
            base = base.sort_values(by="Grand Total", ascending=False)

            # Bottom Grand Total row
            total_row_val = int(base["Grand Total"].sum())
            base = pd.concat([base, pd.DataFrame({"Grand Total": [total_row_val]}, index=["Grand Total"])], axis=0)
            base.index.name = key
            return base

        # -------- Case B: Date columns exist ----------
        # Ensure date coverage + order
        for d in date_cols:
            if d not in tab.columns:
                tab[d] = 0
        tab = tab.reindex(columns=date_cols)

        # Include all labels (to preserve ordering), then fill with zeros
        tab = tab.reindex(label_universe).fillna(0).astype(int)

        # Add per-row Grand Total
        tab["Grand Total"] = tab.sum(axis=1)

        # 🔴 Drop any rows that have Grand Total == 0 (ignore empty/no-count labels)
        tab = tab[tab["Grand Total"] > 0]

        # If everything got dropped, still return a minimal table with zero totals for each date + Grand Total
        if tab.empty:
            total_series = {d: 0 for d in date_cols}
            total_series["Grand Total"] = 0
            out = pd.DataFrame([total_series], index=["Grand Total"])
            out.index.name = key
            return out

        # Sort by Grand Total desc and append bottom Grand Total row
        tab = tab.sort_values(by="Grand Total", ascending=False)
        total_row = tab.sum(axis=0)
        total_row.name = "Grand Total"
        tab = pd.concat([tab, total_row.to_frame().T], axis=0)

        tab.index.name = key
        return tab

    def enforce_date_columns(table_df: pd.DataFrame, date_cols: List[str]) -> pd.DataFrame:
        if not date_cols:
            if "Grand Total" not in table_df.columns:
                table_df["Grand Total"] = table_df.sum(axis=1)
            return table_df[["Grand Total"]]
        for d in date_cols:
            if d not in table_df.columns:
                table_df[d] = 0
        expected = [*date_cols, "Grand Total"]
        if "Grand Total" not in table_df.columns:
            table_df["Grand Total"] = table_df.sum(axis=1)
        return table_df[expected]

    # ----------------------------
    # Key resolution (accept ANY column name; tolerate [__type], spaces, case, NBSP)
    # ----------------------------
    def _strip_brackets_once(s: str) -> str:
        s = s.strip()
        if len(s) >= 2 and ((s[0] == '[' and s[-1] == ']') or (s[0] == '(' and s[-1] == ')') or (s[0] == '{' and s[-1] == '}')):
            return s[1:-1].strip()
        return s

    def _collapse_whitespace(s: str) -> str:
        # Normalize all whitespace (including NBSP) to a single space
        return re.sub(r"[\s\u00A0]+", " ", s).strip()

    def resolve_key_name(df_columns: List[str], requested: str) -> Optional[str]:
        """
        Resolve user-typed key to actual df column name.
        Matching order:
          1) exact
          2) exact on trimmed
          3) exact on bracket-stripped
          4) case-insensitive for 1-3
          5) match against DF columns normalized by stripping ONE layer of brackets
          6) match against DF columns normalized by collapsing whitespace
        """
        cols = list(df_columns)
        req = requested
        req_trim = req.strip()
        req_nobr = _strip_brackets_once(req_trim)
        req_cw = _collapse_whitespace(req_trim)

        # 1-3 exacts
        for cand in (req, req_trim, req_nobr):
            if cand in cols:
                return cand

        # 4) case-insensitive for 1-3
        lower_map = {c.lower(): c for c in cols}
        for cand in (req, req_trim, req_nobr):
            lc = cand.lower()
            if lc in lower_map:
                return lower_map[lc]

        # 5) DF normalized by stripping brackets
        df_nobr_map = {_strip_brackets_once(c): c for c in cols}
        for cand in (req_trim, req_nobr):
            if cand in df_nobr_map:
                return df_nobr_map[cand]
            lc = cand.lower()
            df_nobr_lower = {k.lower(): v for k, v in df_nobr_map.items()}
            if lc in df_nobr_lower:
                return df_nobr_lower[lc]

        # 6) DF normalized by collapsing whitespace
        df_cw_map = {_collapse_whitespace(c): c for c in cols}
        if req_cw in df_cw_map:
            return df_cw_map[req_cw]
        lc = req_cw.lower()
        df_cw_lower = {k.lower(): v for k, v in df_cw_map.items()}
        if lc in df_cw_lower:
            return df_cw_lower[lc]

        return None

    # ----------------------------
    # Styling
    # ----------------------------
    def style_sheet_table(wb_path: Path, sheet_name: str) -> None:
        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            wb.save(wb_path)
            return

        ws = wb[sheet_name]
        last_row = ws.max_row
        last_col = ws.max_column

        ws.freeze_panes = "A2"

        if last_row < 2 or last_col < 1:
            wb.save(wb_path)
            return

        end_cell = f"{get_column_letter(last_col)}{last_row}"

        # Unique table name across workbook
        def _collect_existing_table_names(workbook):
            names = set()
            for w in workbook.worksheets:
                tbls = getattr(w, "tables", None)
                if isinstance(tbls, dict):
                    names.update(tbls.keys())
                elif tbls is not None:
                    for t in tbls:
                        if hasattr(t, "displayName"):
                            names.add(t.displayName)
                        elif isinstance(t, str):
                            names.add(t)
                for t in getattr(w, "_tables", []):
                    if hasattr(t, "displayName"):
                        names.add(t.displayName)
            return names

        existing_table_names = _collect_existing_table_names(wb)
        base_name = re.sub(r"\W+", "_", f"{sheet_name}_Table").strip("_") or "Table1"
        display_name = base_name[:60]
        suffix = 1
        while display_name in existing_table_names:
            suffix += 1
            display_name = (f"{base_name}_{suffix}")[:60]

        table = Table(displayName=display_name, ref=f"A1:{end_cell}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Wrap + vertical top alignment
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto width (scan up to 500 rows)
        for col_idx in range(1, last_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min(last_row, 500) + 1):
                cell_val = ws[f"{col_letter}{row_idx}"].value
                if cell_val is None:
                    continue
                max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)

        # Bold final row if 'Grand Total' row
        try:
            for cell in ws[last_row]:
                if isinstance(cell.value, (int, float)) or cell.value == "Grand Total":
                    cell.font = Font(bold=True)
        except Exception:
            pass

        wb.save(wb_path)

    def fixed_output_name() -> Path:
        return Path("sub_dar_individual_keys_datacount.xlsx")

    # ----------------------------
    # Execute
    # ----------------------------
    try:
        input_path, requested_keys = prompt_for_file_and_keys()
        print(f"📥 Input file: {input_path.resolve()}")

        # Load sheet (preserve 'null' strings and empty strings)
        df = read_sheet(input_path)

        # Column audit to reveal hidden characters
        print("🧭 Columns found in sheet (repr):")
        for c in df.columns:
            print("   -", repr(c))

        # DateOnly (if present)
        df, date_cols = add_date_only_if_present(df)

        if date_cols:
            print(f"📅 Included dates: {len(date_cols)} distinct days ({date_cols[0]} → {date_cols[-1]})")
        else:
            print("📅 No valid 'Date' column or no valid dates; will produce 'Grand Total' only.")

        print(f"🔑 Keys (requested): {', '.join(requested_keys)}")

        # Resolve keys, ALWAYS create a sheet
        df_cols = list(df.columns)
        resolved: List[tuple[str, str]] = []  # (requested, actual-in-df)
        created_placeholders: List[str] = []

        for req_key in requested_keys:
            actual = resolve_key_name(df_cols, req_key)
            if actual is None:
                placeholder_col = req_key  # create placeholder so sheet is produced
                if placeholder_col not in df.columns:
                    df[placeholder_col] = pd.NA
                    df_cols.append(placeholder_col)
                    created_placeholders.append(req_key)
                resolved.append((req_key, placeholder_col))
            else:
                resolved.append((req_key, actual))

        print("🔎 Key resolution:")
        for req, act in resolved:
            note = " (placeholder)" if req in created_placeholders else ""
            print(f"   - {repr(req)} -> {repr(act)}{note}")

        out_path = fixed_output_name()

        total_steps = len(resolved) * 2  # write + style
        step = 0
        print_progress(step, total_steps)

        used_sheet_names = set()

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for (requested_name, actual_col) in resolved:
                table_df = build_table_for_key(df, actual_col, date_cols)
                table_df = enforce_date_columns(table_df, date_cols)

                base_name = sanitize_sheet_name(requested_name) or "Sheet1"
                sheet_name = base_name
                suffix = 1
                existing = used_sheet_names | set(writer.sheets.keys())
                if hasattr(writer, "book"):
                    existing |= {ws.title for ws in writer.book.worksheets}
                while sheet_name in existing:
                    suffix += 1
                    sheet_name = (base_name[:28] + f"_{suffix}")[:31]

                table_df.to_excel(writer, index=True, sheet_name=sheet_name, index_label=requested_name)
                used_sheet_names.add(sheet_name)

                step += 1
                print_progress(step, total_steps)

        # Style sheets
        for sheet_name in used_sheet_names:
            try:
                style_sheet_table(out_path, sheet_name)
            except Exception as se:
                print(f"\n⚠️ Styling skipped for sheet '{sheet_name}': {se}")
            step += 1
            print_progress(step, total_steps)

        print(f"\n✅ Done! Wrote analysis to '{out_path.name}'.")
        print(f"📄 Output path: {out_path.resolve()}")

        if created_placeholders:
            print("\nℹ️ Note: The following requested keys were not found as columns;")
            print("   placeholder sheets were created (all zeros) so you can see them anyway:")
            for k in created_placeholders:
                print("   -", k)

    except Exception as e:
        print("❌ Failed to run individual_keys_datacount.")
        print("Error:", str(e))
        import traceback
        traceback.print_exc()
        sys.exit(2)


# if __name__ == "__main__":
#     sub_dar_individual_keys_datacount()










# def individual_keys_datacount():
#     """
#     Build per-key, per-day count tables from a parsed logs Excel and
#     save them into a new Excel file (one sheet per key) with table styling.

#     Implemented:
#       - Default input file on Enter: 'sub_dar_refined.xlsx' in the current directory.
#       - Use 'Date' column; consider only the date portion (YYYY-MM-DD), ignore time.
#       - Do NOT delete old output files.
#       - Preserve values exactly (no trimming/normalizing):
#           * Display & grouping mapping:
#               - NaN -> literal "null"
#               - Non-null -> str(value) (so "" stays "", "   " stays as spaces, numbers -> "123", etc.)
#       - Count rows only when 'Date' parsed (non-null DateOnly).
#       - Every distinct value appearing in the column (even only with null Date) is included as a row
#         (with zeros if it never occurs with a valid Date).
#       - Output overwrites 'individual_keys_datacount.xlsx'.
#       - Silences pandas FutureWarning by passing observed=False in groupby.

#     For each key:
#       - Columns: all distinct dates (YYYY-MM-DD) in chronological ASC order, then 'Grand Total'.
#       - Rows: distinct display values (as-is, with NaN -> "null"), sorted by 'Grand Total' (desc),
#               plus a final 'Grand Total' row.
#     """

#     # ----------------------------
#     # Imports & setup
#     # ----------------------------
#     import sys
#     import re
#     from datetime import datetime
#     from pathlib import Path
#     import pandas as pd
#     from openpyxl import load_workbook
#     from openpyxl.worksheet.table import Table, TableStyleInfo
#     from openpyxl.utils import get_column_letter
#     from openpyxl.styles import Alignment, Font

#     # ----------------------------
#     # Progress Bar
#     # ----------------------------
#     def print_progress(current: int, total: int, bar_width: int = 40) -> None:
#         """Render a single overall progress bar (0..100%)."""
#         if total <= 0:
#             return
#         ratio = max(0.0, min(1.0, current / total))
#         filled = int(bar_width * ratio)
#         bar = '#' * filled + '-' * (bar_width - filled)
#         percent = int(ratio * 100)
#         sys.stdout.write(f"\rProcessing: [{bar}] {percent}%")
#         sys.stdout.flush()
#         if current >= total:
#             sys.stdout.write("\n")

#     # ----------------------------
#     # Helpers (sheet name, input prompts)
#     # ----------------------------
#     def sanitize_sheet_name(name: str) -> str:
#         """Remove invalid Excel sheet name chars and cap length to 31."""
#         invalid = r'[:\\/\?\*\[\]]'
#         clean = re.sub(invalid, "_", name or "Sheet1")
#         return clean[:31] if clean else "Sheet1"

#     def prompt_for_file_and_keys() -> tuple[Path, list[str]]:
#         """
#         Prompt for the Excel file path (Enter -> use 'sub_dar_refined.xlsx' in current dir),
#         then prompt for keys (comma/space separated).
#         """
#         file_in = input("📄 Enter Excel path (or press Enter to use 'sub_dar_refined.xlsx' in current folder): ").strip()
#         if file_in:
#             xlsx_path = Path(file_in)
#         else:
#             xlsx_path = Path("sub_dar_refined.xlsx")

#         if not xlsx_path.exists():
#             raise FileNotFoundError(
#                 f"Input file not found: {xlsx_path}\n"
#                 "Provide a valid path or place 'sub_dar_refined.xlsx' in the current directory."
#             )

#         keys_raw = input("🔑 Enter keys (comma or space separated): ").strip()
#         if not keys_raw:
#             raise ValueError("No keys provided. Please enter at least one key name.")
#         parts = re.split(r"[,\s]+", keys_raw)
#         keys: list[str] = []
#         for p in parts:
#             if p and p not in keys:
#                 keys.append(p)
#         return xlsx_path, keys

#     # ----------------------------
#     # Data I/O helpers
#     # ----------------------------
#     def read_sheet(xlsx_path: Path) -> pd.DataFrame:
#         """
#         Read 'TopKeys' sheet if present, else the first sheet.
#         Do NOT force dtype; preserve native types & nulls.
#         """
#         xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
#         sheet = "TopKeys" if "TopKeys" in xls.sheet_names else xls.sheet_names[0]
#         df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
#         df.columns = [str(c).strip() for c in df.columns]
#         return df

#     def add_date_only_column(df: pd.DataFrame) -> pd.DataFrame:
#         """
#         Create a 'DateOnly' column from 'Date', formatted as 'YYYY-MM-DD' (date part only).
#         NaT -> NaN in the result.
#         """
#         if "Date" not in df.columns:
#             raise ValueError("Input sheet missing 'Date' column.")
#         ts = pd.to_datetime(df["Date"], errors="coerce")
#         df["DateOnly"] = ts.dt.strftime("%Y-%m-%d")  # NaT -> NaN
#         return df

#     def chronological_dates(df: pd.DataFrame) -> list[str]:
#         """
#         All distinct non-null 'YYYY-MM-DD' dates in chronological ASC order (unique).
#         """
#         vals = df["DateOnly"].dropna().astype(str)
#         if vals.empty:
#             return []
#         dt = pd.to_datetime(vals, format="%Y-%m-%d", errors="coerce").dropna()
#         unique_sorted = pd.DatetimeIndex(pd.unique(dt)).sort_values()
#         return unique_sorted.strftime("%Y-%m-%d").tolist()

#     # ----------------------------
#     # Display mapping (as-is, with NaN -> "null")
#     # ----------------------------
#     def to_label(value) -> str:
#         """
#         Return the display label for a value:
#           - NaN -> "null"
#           - else -> str(value) (no trimming or case change)
#         """
#         if pd.isna(value):
#             return "null"
#         return str(value)

#     def labels_series(raw_series: pd.Series) -> pd.Series:
#         """
#         Build the label series (string) for a column using to_label for each cell.
#         """
#         return raw_series.map(to_label)

#     def first_seen_labels(raw_series: pd.Series) -> list[str]:
#         """
#         Unique labels in first-appearance order (using to_label), preserving exact text
#         (including empty "", whitespace, numbers-as-text, etc.).
#         """
#         order: list[str] = []
#         for v in raw_series:
#             lab = to_label(v)
#             if lab not in order:
#                 order.append(lab)
#         return order

#     # ----------------------------
#     # Counting utilities
#     # ----------------------------
#     def counts_table(df: pd.DataFrame, key: str) -> pd.DataFrame:
#         """
#         Compute counts per (label, DateOnly) using groupby.
#         - Label uses to_label (NaN -> "null", others str(value) exactly).
#         - Only rows with non-null DateOnly contribute to counts.
#         - Pass observed=False to silence pandas FutureWarning and retain current behavior.
#         Returns a DataFrame indexed by label (string), columns = DateOnly.
#         """
#         work = df[[key, "DateOnly"]].copy()
#         work["__label__"] = labels_series(work[key])
#         work = work[work["DateOnly"].notna()]
#         if work.empty:
#             return pd.DataFrame()
#         tab = (
#             work.groupby(["__label__", "DateOnly"], observed=False)  # <- key change
#                 .size()
#                 .unstack("DateOnly", fill_value=0)
#         )
#         tab.index.name = key
#         return tab

#     def build_table_for_key(df: pd.DataFrame, key: str, date_cols: list[str]) -> pd.DataFrame:
#         """
#         Build the final table for a given key:
#         - Columns: date_cols (chronological ASC), then 'Grand Total'
#         - Rows: ALL distinct labels from the column (as-is + "null" for NaNs), even if they never
#           occur with a valid date (they appear with 0s), sorted by 'Grand Total' desc,
#           then a final 'Grand Total' row.
#         """
#         if key not in df.columns:
#             raise KeyError(f"Key '{key}' not found in sheet columns.")

#         # Universe of labels (as-is, incl. "null")
#         label_universe = first_seen_labels(df[key])

#         # Counts for labels that appear with valid DateOnly
#         tab = counts_table(df, key)

#         # If there are dates, ensure all date columns exist in order
#         if tab.empty:
#             # Start empty frame with all labels, no date cols
#             base = pd.DataFrame(index=label_universe)
#             base.index.name = key
#             for d in date_cols:
#                 base[d] = 0
#             # Add Grand Total
#             base["Grand Total"] = 0
#             # If no labels at all, keep a minimal table with just the grand total row
#             if base.empty:
#                 base = pd.DataFrame({"Grand Total": [0]})
#                 base.index = ["Grand Total"]
#                 base.index.name = key
#                 return base
#             # Add overall Grand Total row at bottom
#             grand_row = base.sum(axis=0)
#             grand_row.name = "Grand Total"
#             base = pd.concat([base, grand_row.to_frame().T], axis=0)
#             return base

#         # Ensure all date columns in desired order; add missing as 0
#         for d in date_cols:
#             if d not in tab.columns:
#                 tab[d] = 0
#         tab = tab.reindex(columns=date_cols)

#         # Reindex rows to include every label (even those with only null Dates)
#         tab = tab.reindex(label_universe).fillna(0).astype(int)

#         # Add per-row Grand Total
#         tab["Grand Total"] = tab.sum(axis=1)

#         # Sort rows by Grand Total desc (ties resolved by first-seen order already in index)
#         tab = tab.sort_values(by="Grand Total", ascending=False)

#         # Add Grand Total row at bottom
#         total_row = tab.sum(axis=0)
#         total_row.name = "Grand Total"
#         tab = pd.concat([tab, total_row.to_frame().T], axis=0)

#         tab.index.name = key
#         return tab

#     def enforce_date_columns(table_df: pd.DataFrame, date_cols: list[str]) -> pd.DataFrame:
#         """
#         Strictly enforce presence and order of date columns for the table.
#         Adds zero-filled columns for any missing dates and reorders columns.
#         Handles the case where date_cols is empty (then only 'Grand Total' remains).
#         """
#         for d in date_cols:
#             if d not in table_df.columns:
#                 table_df[d] = 0
#         expected = [*date_cols, "Grand Total"]
#         if "Grand Total" not in table_df.columns:
#             table_df["Grand Total"] = table_df.sum(axis=1)
#         return table_df[expected]

#     # ----------------------------
#     # Styling
#     # ----------------------------
#     def style_sheet_table(wb_path: Path, sheet_name: str) -> None:
#         """Add an Excel table, freeze header row, wrap cell text, and auto-fit column widths."""
#         wb = load_workbook(wb_path)
#         if sheet_name not in wb.sheetnames:
#             wb.save(wb_path)
#             return

#         ws = wb[sheet_name]
#         last_row = ws.max_row
#         last_col = ws.max_column

#         # Freeze header row (A2 keeps header visible)
#         ws.freeze_panes = "A2"

#         # If there's only a header, skip table creation to avoid openpyxl errors
#         if last_row < 2 or last_col < 1:
#             wb.save(wb_path)
#             return

#         end_cell = f"{get_column_letter(last_col)}{last_row}"

#         # Ensure unique table display name across workbook (robust across openpyxl versions)
#         def _collect_existing_table_names(workbook) -> set[str]:
#             names = set()
#             for w in workbook.worksheets:
#                 tbls = getattr(w, "tables", None)
#                 if isinstance(tbls, dict):         # openpyxl 3.1+: dict {name: Table}
#                     names.update(tbls.keys())
#                 elif tbls is not None:              # older shapes: iterable of Table or names
#                     for t in tbls:
#                         if hasattr(t, "displayName"):
#                             names.add(t.displayName)
#                         elif isinstance(t, str):
#                             names.add(t)
#                 for t in getattr(w, "_tables", []): # very old: private _tables list
#                     if hasattr(t, "displayName"):
#                         names.add(t.displayName)
#             return names

#         existing_table_names = _collect_existing_table_names(wb)
#         base_name = re.sub(r"\W+", "_", f"{sheet_name}_Table").strip("_") or "Table1"
#         display_name = base_name[:60]
#         suffix = 1
#         while display_name in existing_table_names:
#             suffix += 1
#             display_name = (f"{base_name}_{suffix}")[:60]

#         table = Table(displayName=display_name, ref=f"A1:{end_cell}")
#         style = TableStyleInfo(
#             name="TableStyleMedium9",
#             showFirstColumn=False,
#             showLastColumn=False,
#             showRowStripes=True,
#             showColumnStripes=False,
#         )
#         table.tableStyleInfo = style
#         ws.add_table(table)

#         # Wrap + vertical top alignment (data rows)
#         for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
#             for cell in row:
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#         # Auto width (capped, scanning up to 500 rows for performance)
#         for col_idx in range(1, last_col + 1):
#             col_letter = get_column_letter(col_idx)
#             max_len = 0
#             for row_idx in range(1, min(last_row, 500) + 1):
#                 cell_val = ws[f"{col_letter}{row_idx}"].value
#                 if cell_val is None:
#                     continue
#                 max_len = max(max_len, len(str(cell_val)))
#             ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)

#         # Bold final 'Grand Total' row if present
#         try:
#             for cell in ws[last_row]:
#                 if isinstance(cell.value, (int, float)) or cell.value == "Grand Total":
#                     cell.font = Font(bold=True)
#         except Exception:
#             pass

#         wb.save(wb_path)

#     def fixed_output_name() -> Path:
#         """Return the fixed output filename (no timestamp) so runs overwrite the same file."""
#         return Path("individual_keys_datacount.xlsx")

#     # ----------------------------
#     # Execute
#     # ----------------------------
#     try:
#         # Prompt for file and keys
#         input_path, keys = prompt_for_file_and_keys()
#         print(f"📥 Input file: {input_path.resolve()}")
#         print(f"🔑 Keys: {', '.join(keys)}")

#         # Load data and derive date-only column
#         df = read_sheet(input_path)
#         df = add_date_only_column(df)

#         # Chronological ASC dates (YYYY-MM-DD), unique
#         date_cols = chronological_dates(df)
#         # Ensure uniqueness, preserve order
#         date_cols = list(dict.fromkeys(date_cols))

#         # Lock DateOnly categorical order if there are any dates
#         if date_cols:
#             df["DateOnly"] = pd.Categorical(df["DateOnly"], categories=date_cols, ordered=True)

#         # Pre-validate keys
#         missing = [k for k in keys if k not in df.columns]
#         present = [k for k in keys if k in df.columns]
#         if missing:
#             print(f"⚠️ Skipping missing keys: {', '.join(missing)}")
#         if not present:
#             raise KeyError("None of the provided keys are present in the input sheet.")

#         # Prepare output (overwrite same file each run)
#         out_path = fixed_output_name()

#         total_steps = len(present) + len(present)  # write + style
#         step = 0
#         print_progress(step, total_steps)

#         # Ensure unique sheet names
#         used_sheet_names = set()

#         # Write one sheet per key with strict date coverage
#         with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
#             for key in present:
#                 table_df = build_table_for_key(df, key, date_cols)
#                 table_df = enforce_date_columns(table_df, date_cols)

#                 base_name = sanitize_sheet_name(key) or "Sheet1"
#                 sheet_name = base_name
#                 suffix = 1
#                 existing = used_sheet_names | set(writer.sheets.keys())
#                 if hasattr(writer, "book"):
#                     existing |= {ws.title for ws in writer.book.worksheets}
#                 while sheet_name in existing:
#                     suffix += 1
#                     sheet_name = (base_name[:28] + f"_{suffix}")[:31]

#                 table_df.to_excel(writer, index=True, sheet_name=sheet_name, index_label=key)
#                 used_sheet_names.add(sheet_name)

#                 step += 1
#                 print_progress(step, total_steps)

#         # Style each sheet (table, wrap, widths)
#         for sheet_name in used_sheet_names:
#             try:
#                 style_sheet_table(out_path, sheet_name)
#             except Exception as se:
#                 print(f"\n⚠️ Styling skipped for sheet '{sheet_name}': {se}")
#             step += 1
#             print_progress(step, total_steps)

#         # Friendly date summary
#         if date_cols:
#             print(f"📅 Included dates: {len(date_cols)} distinct days ({date_cols[0]} → {date_cols[-1]})")
#         else:
#             print("📅 No valid dates found; rows will have Grand Total = 0.")

#         print(f"\n✅ Done! Wrote analysis to '{out_path.name}'.")
#         print(f"📄 Output path: {out_path.resolve()}")

#     except Exception as e:
#         print("❌ Failed to run individual_keys_datacount.")
#         print("Error:", str(e))
#         import traceback
#         traceback.print_exc()
#         sys.exit(2)






# def individual_keys_datacount():
#     """
#     Build per-key, per-day count tables from a parsed logs Excel and
#     save them into a new Excel file (one sheet per key) with table styling.

#     - Prompts for parsed logs Excel path (Enter => auto-picks latest 'parsed_logs_*.xlsx').
#     - Prompts for keys (comma/space separated). 'S.NO' and 'Timestamp' are implicit defaults; you only provide keys.
#     - Robust Timestamp parsing; adds 'Unknown Date' for rows that fail to parse (can be disabled).
#     - For each key:
#         - Columns: ALL dates present (from 'Timestamp', formatted dd/MM/YYYY) in chronological ASC order,
#                    followed by 'Unknown Date' (if present), then 'Grand Total'.
#         - Rows: distinct values of the key (cleaned/stripped), sorted by 'Grand Total' (desc),
#                 followed by a 'Grand Total' row.
#     - Writes 'individual_keys_datacount_YYYYMMDD_HHMMSS.xlsx' with one sheet per key.
#     """
#     # ----------------------------
#     # Local imports & setup
#     # ----------------------------
#     import sys
#     import re
#     from datetime import datetime
#     from pathlib import Path
#     import pandas as pd
#     from openpyxl import load_workbook
#     from openpyxl.worksheet.table import Table, TableStyleInfo
#     from openpyxl.utils import get_column_letter
#     from openpyxl.styles import Alignment
#     import os

#     INCLUDE_UNKNOWN_DATE = True  # Count rows with unparseable Timestamp under 'Unknown Date'

#     # ----------------------------
#     # Helpers
#     # ----------------------------
#     def sanitize_sheet_name(name: str) -> str:
#         """Remove invalid Excel sheet name chars and cap length to 31."""
#         invalid = r'[:\\/\?\*\[\]]'
#         clean = re.sub(invalid, "_", name or "Sheet1")
#         return clean[:31] if clean else "Sheet1"

#     def find_latest_parsed_logs() -> Path:
#         """Pick latest parsed_logs_YYYYMMDD_HHMMSS.xlsx by timestamp in filename; fallback to mtime."""
#         cwd = Path.cwd()
#         candidates = list(cwd.glob("parsed_logs_*.xlsx"))
#         if not candidates:
#             raise FileNotFoundError("No 'parsed_logs_*.xlsx' files found in the current directory.")

#         def sort_key(p: Path):
#             m = re.match(r"parsed_logs_(\d{8})_(\d{6})\.xlsx$", p.name)
#             if m:
#                 ymd, hms = m.group(1), m.group(2)
#                 try:
#                     return datetime.strptime(f"{ymd}_{hms}", "%Y%m%d_%H%M%S")
#                 except Exception:
#                     return datetime.fromtimestamp(p.stat().st_mtime)
#             return datetime.fromtimestamp(p.stat().st_mtime)

#         candidates.sort(key=sort_key, reverse=True)
#         return candidates[0]

#     def prompt_for_file_and_keys() -> tuple[Path, list[str]]:
#         """
#         Prompt for the Excel file path (Enter -> auto-pick latest),
#         then prompt for keys (comma/space separated).
#         """
#         file_in = input("📄 Enter parsed logs Excel path (or press Enter to auto-pick latest 'parsed_logs_*.xlsx'): ").strip()
#         if file_in:
#             xlsx_path = Path(file_in)
#             if not xlsx_path.exists():
#                 raise FileNotFoundError(f"File not found: {xlsx_path}")
#         else:
#             xlsx_path = find_latest_parsed_logs()

#         keys_raw = input("🔑 Enter keys (comma or space separated): ").strip()
#         if not keys_raw:
#             raise ValueError("No keys provided. Please enter at least one key name.")
#         parts = re.split(r"[,\s]+", keys_raw)
#         keys: list[str] = []
#         for p in parts:
#             if p and p not in keys:
#                 keys.append(p)
#         return xlsx_path, keys

#     def read_sheet(xlsx_path: Path) -> pd.DataFrame:
#         """Read 'TopKeys' sheet if present, else the first sheet, as strings."""
#         xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
#         sheet = "TopKeys" if "TopKeys" in xls.sheet_names else xls.sheet_names[0]
#         df = pd.read_excel(xls, sheet_name=sheet, dtype=str, engine="openpyxl")
#         df.columns = [c.strip() for c in df.columns]
#         return df

#     def clean_timestamp_series(ts_series: pd.Series) -> pd.Series:
#         """
#         Robust clean of Timestamp strings:
#         - strip, remove control chars
#         - replace 'T' with ' ' (ISO)
#         - strip trailing 'Z'
#         - strip timezone offsets (+05:30, -0700) so day-level grouping works
#         """
#         s = ts_series.astype(str)
#         s = s.str.replace(r"[\r\n\t]", "", regex=True).str.strip()
#         s = s.str.replace("T", " ")
#         s = s.str.replace("Z", "", regex=False)
#         # Remove offsets like '+05:30', '-0700' at the end
#         s = s.str.replace(r"([+\-]\d{2}:\d{2})$", "", regex=True)
#         s = s.str.replace(r"([+\-]\d{4})$", "", regex=True)
#         return s

#     def add_date_column(df: pd.DataFrame) -> pd.DataFrame:
#         """
#         Create a 'DateStr' column from 'Timestamp', formatted as dd/MM/YYYY.
#         """
#         if "Timestamp" not in df.columns:
#             raise ValueError("Input sheet missing 'Timestamp' column.")
#         cleaned = clean_timestamp_series(df["Timestamp"])
#         # Parse flexibly; coerce failures
#         ts = pd.to_datetime(cleaned, errors="coerce", utc=False)
#         df["DateStr"] = ts.dt.strftime("%d/%m/%Y")
#         if INCLUDE_UNKNOWN_DATE:
#             df.loc[df["DateStr"].isna(), "DateStr"] = "Unknown Date"
#         return df

#     def chronological_dates(df: pd.DataFrame) -> list[str]:
#         """
#         Return ALL dates present in the file in chronological ASC order.
#         'Unknown Date' (if any) is appended at the end.
#         """
#         vals = df["DateStr"].dropna().astype(str)
#         normal_dates = vals[vals != "Unknown Date"]
#         unknown_present = (vals == "Unknown Date").any()

#         parsed = pd.to_datetime(normal_dates, format="%d/%m/%Y", errors="coerce")
#         sorted_unique = sorted(set(d.strftime("%d/%m/%Y") for d in parsed.dropna()))
#         if unknown_present:
#             sorted_unique.append("Unknown Date")
#         return sorted_unique

#     def clean_key_series(series: pd.Series) -> pd.Series:
#         """
#         Clean a key series for grouping:
#         - cast to str
#         - strip whitespace
#         - convert explicit text 'null'/'None'/'nan' to empty
#         """
#         s = series.astype(str).str.strip()
#         s_lower = s.str.lower()
#         s = s.where(~s_lower.isin({"null", "none", "nan"}), "")
#         return s

#     def first_seen_order(values: pd.Series) -> list[str]:
#         """Unique values in first-appearance order (ignoring blanks)."""
#         order: list[str] = []
#         for v in values.astype(str):
#             s = v.strip()
#             if not s:
#                 continue
#             if s not in order:
#                 order.append(s)
#         return order

#     def crosstab_counts(df: pd.DataFrame, key: str) -> pd.DataFrame:
#         """
#         Generate counts using crosstab to avoid pivot quirks.
#         Returns a DataFrame indexed by cleaned key values, columns = DateStr.
#         """
#         work = df[[key, "DateStr"]].copy()
#         work[key] = clean_key_series(work[key])
#         # Only non-empty values contribute
#         work = work[work[key] != ""]
#         if work.empty:
#             return pd.DataFrame()
#         tab = pd.crosstab(work[key], work["DateStr"])
#         return tab

#     def build_table_for_key(df: pd.DataFrame, key: str, date_cols: list[str]) -> pd.DataFrame:
#         """
#         Build the final table for a given key:
#         - Columns: date_cols (chronological ASC), then 'Grand Total'
#         - Rows: distinct cleaned values (first-seen order), sorted by Grand Total (desc), then Grand Total row
#         """
#         if key not in df.columns:
#             raise KeyError(f"Key '{key}' not found in sheet columns.")

#         # Build crosstab counts
#         tab = crosstab_counts(df, key)

#         # Establish row order from first-seen appearance across entire dataset for this key
#         cat_order = first_seen_order(clean_key_series(df[key]))

#         # If tab is empty (no non-empty values), return a minimal table (safe for Excel table creation)
#         if tab.empty or not cat_order:
#             base = pd.DataFrame({d: [0] for d in date_cols})
#             base["Grand Total"] = 0
#             base.index = ["Grand Total"]
#             base.index.name = key
#             return base

#         # Ensure all date columns exist in ASC order (fill zeros where missing)
#         for d in date_cols:
#             if d not in tab.columns:
#                 tab[d] = 0
#         tab = tab[date_cols]

#         # Reindex rows to first-seen order, fill zeros for missing
#         tab = tab.reindex(cat_order).fillna(0).astype(int)

#         # Add per-row Grand Total
#         tab["Grand Total"] = tab.sum(axis=1)

#         # Sort rows by Grand Total desc (ties broken by first-seen order)
#         row_totals = tab["Grand Total"].copy()
#         orig_row_index = {cat: i for i, cat in enumerate(cat_order)}
#         sorted_rows = sorted(cat_order, key=lambda r: (-row_totals.loc[r], orig_row_index[r]))
#         tab = tab.loc[sorted_rows]

#         # Add Grand Total row at bottom
#         total_row = tab.sum(axis=0)
#         total_row.name = "Grand Total"
#         tab = pd.concat([tab, total_row.to_frame().T], axis=0)

#         # Set index name
#         tab.index.name = key
#         return tab

#     def enforce_date_columns(table_df: pd.DataFrame, date_cols: list[str]) -> pd.DataFrame:
#         """
#         Strictly enforce presence and order of date columns for the table.
#         Adds zero-filled columns for any missing dates and reorders columns.
#         """
#         for d in date_cols:
#             if d not in table_df.columns:
#                 table_df[d] = 0
#         expected = [*date_cols, "Grand Total"]
#         # If Grand Total missing (shouldn't happen), add it
#         if "Grand Total" not in table_df.columns:
#             table_df["Grand Total"] = table_df.sum(axis=1)
#         return table_df[expected]

#     def style_sheet_table(wb_path: Path, sheet_name: str) -> None:
#         """Add an Excel table, freeze header row, wrap cell text, and auto-fit column widths."""
#         wb = load_workbook(wb_path)
#         if sheet_name not in wb.sheetnames:
#             wb.save(wb_path)
#             return

#         ws = wb[sheet_name]
#         last_row = ws.max_row
#         last_col = ws.max_column

#         # Freeze header row regardless
#         ws.freeze_panes = "A2"

#         # If there's only a header, skip table creation to avoid openpyxl errors
#         if last_row < 2 or last_col < 1:
#             for col_idx in range(1, last_col + 1):
#                 col_letter = get_column_letter(col_idx)
#                 ws.column_dimensions[col_letter].width = min(20, 80)
#             wb.save(wb_path)
#             return

#         end_cell = f"{get_column_letter(last_col)}{last_row}"
#         table_name = re.sub(r"\W+", "_", f"{sheet_name}_Table")[:31]
#         table = Table(displayName=table_name, ref=f"A1:{end_cell}")
#         style = TableStyleInfo(
#             name="TableStyleMedium9",
#             showFirstColumn=False,
#             showLastColumn=False,
#             showRowStripes=True,
#             showColumnStripes=False,
#         )
#         table.tableStyleInfo = style
#         ws.add_table(table)

#         # Wrap + vertical top alignment
#         for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
#             for cell in row:
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#         # Auto width (capped, scanning up to 500 rows for performance)
#         for col_idx in range(1, last_col + 1):
#             col_letter = get_column_letter(col_idx)
#             max_len = 0
#             for row_idx in range(1, min(last_row, 500) + 1):
#                 cell_val = ws[f"{col_letter}{row_idx}"].value
#                 if cell_val is None:
#                     continue
#                 max_len = max(max_len, len(str(cell_val)))
#             ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

#         wb.save(wb_path)

#     def timestamped_output_name(prefix: str = "individual_keys_datacount", ext: str = ".xlsx") -> Path:
#         """Generate a filename like: individual_keys_datacount_YYYYMMDD_HHMMSS.xlsx using local time."""
#         ts = datetime.now().strftime("%Y%m%d_%H%M%S")
#         return Path(f"{prefix}_{ts}{ext}")

#     # ----------------------------
#     # Execute
#     # ----------------------------
#     try:
#         # Prompt for file and keys
#         input_path, keys = prompt_for_file_and_keys()
#         print(f"📥 Input file: {input_path.resolve()}")
#         print(f"🔑 Keys: {', '.join(keys)}")

#         # Load data and derive dates
#         df = read_sheet(input_path)
#         df = add_date_column(df)

#         # Chronological ASC dates (all present), Unknown Date (if any) appended at end
#         date_cols = chronological_dates(df)
#         if not date_cols:
#             if INCLUDE_UNKNOWN_DATE and (df["DateStr"] == "Unknown Date").any():
#                 date_cols = ["Unknown Date"]
#             else:
#                 raise ValueError("No valid dates parsed from 'Timestamp' column.")

#         # Lock DateStr order and improve memory use
#         df["DateStr"] = pd.Categorical(df["DateStr"], categories=date_cols, ordered=True)

#         # Prepare output
#         out_path = timestamped_output_name(prefix="individual_keys_datacount", ext=".xlsx")
#         if out_path.exists():
#             try:
#                 out_path.unlink()
#             except Exception:
#                 pass

#         # Pre-validate keys
#         missing = [k for k in keys if k not in df.columns]
#         present = [k for k in keys if k in df.columns]
#         if missing:
#             print(f"⚠️ Skipping missing keys: {', '.join(missing)}")
#         if not present:
#             raise KeyError("None of the provided keys are present in the input sheet.")

#         # Write one sheet per key with strict date coverage
#         with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
#             for key in present:
#                 table_df = build_table_for_key(df, key, date_cols)

#                 # Strictly enforce date column presence & order (NO DATE SKIPPED)
#                 table_df = enforce_date_columns(table_df, date_cols)

#                 sheet_name = sanitize_sheet_name(key)
#                 table_df.to_excel(writer, index=True, sheet_name=sheet_name)

#         # Style each sheet
#         for key in present:
#             sheet_name = sanitize_sheet_name(key)
#             style_sheet_table(out_path, sheet_name)

#         # Summary for Unknown Date
#         if INCLUDE_UNKNOWN_DATE:
#             unknown_count = (df["DateStr"] == "Unknown Date").sum()
#             total_rows = len(df)
#             if unknown_count:
#                 pct = (unknown_count / total_rows) * 100 if total_rows else 0
#                 print(f"ℹ️ Unknown Date rows: {unknown_count} / {total_rows} ({pct:.2f}%)")

#         # Friendly date summary
#         if date_cols:
#             if "Unknown Date" in date_cols:
#                 core_dates = [d for d in date_cols if d != "Unknown Date"]
#                 range_str = f"{core_dates[0]} → {core_dates[-1]}" if core_dates else "N/A"
#                 print(f"📅 Included dates: {len(core_dates)} distinct days ({range_str}) + 'Unknown Date'")
#             else:
#                 print(f"📅 Included dates: {len(date_cols)} distinct days ({date_cols[0]} → {date_cols[-1]})")

#         print(f"\n✅ Done! Wrote analysis to '{out_path.name}'.")
#         print(f"📄 Output path: {out_path.resolve()}")

#     except Exception as e:
#         print("❌ Failed to run individual_keys_datacount_analysis.")
#         print("Error:", str(e))
#         import traceback
#         traceback.print_exc()
#         sys.exit(2)
